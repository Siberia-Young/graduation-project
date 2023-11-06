from selenium import webdriver
from openpyxl import Workbook
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from io import BytesIO
import os
import requests
import urllib.parse
import time
import random
import datetime
import json

min_delay = 3  # 最小延迟时间（单位：秒）
max_delay = 5  # 最大延迟时间（单位：秒）

row_height = 40
column_width = 14

def login(driver):
    print('登录')
    # driver.get('https://mms.pinduoduo.com/login/sso?platform=wholesale&redirectUrl=https%3A%2F%2Fpifa.pinduoduo.com%2Fclint%2Fapi%2Flogin%3Fredirect%3Dhttps%253A%252F%252Fpifa.pinduoduo.com%252F')
    time.sleep(60)

def scrape_multiple_pages(keyword, start_page, end_page):
    # options = webdriver.FirefoxOptions()
    # driver = webdriver.Remote(
    #     command_executor="http://127.0.0.1:4444", options=options)
    
    options = webdriver.EdgeOptions()
    driver = webdriver.Edge(options=options)

    # 创建带有Selenium Wire的Firefox WebDriver对象
    # options = webdriver.FirefoxOptions()
    # options.set_preference('network.proxy.type', 1)
    # options.set_preference('network.proxy.http', 'localhost')
    # options.set_preference('network.proxy.http_port', 8888)
    # driver = webdriver.Firefox(options=options)

    login(driver)
    print("登录成功")

    # 获取当前时间
    current_time = datetime.datetime.now()
    # 格式化时间字符串
    time_string = current_time.strftime("%Y-%m-%d_%H-%M-%S")
    # 构建文件名
    file_name = f"data/wph/唯品会_{urllib.parse.unquote(keyword)}_{time_string}.xlsx"
    total_num = 0
    record_num = 0

    workbook = Workbook()
    sheet = workbook.active
    headers = ['序号', '电商平台', '关键词', '店铺名称', '店铺网址', '店铺经营主体信息', '商品图片', '商品标题', '商品品牌', '商品链接', '单价', '销售量', '商品评论数', '销售额']
    
    sheet.append(headers)
    for index, cell in enumerate(sheet[1], start=1):
        if index == 16:
            cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
            cell.font = Font(bold=True, color="FFFFFF")
        else:
            cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            cell.font = Font(bold=True, color="000000")
        cell.alignment = Alignment(horizontal='center', vertical='center')
    workbook.save(file_name)

    searchUrl = "https://category.vip.com/suggest.php?keyword="+keyword+"&page=1"
    driver.get(searchUrl)
    html = driver.execute_script(
        "return document.documentElement.outerHTML")
    # 创建 Beautiful Soup 对象
    soup = BeautifulSoup(html, "html.parser")

    # 使用 select 方法查找指定的元素
    try:
        # elements = soup.select('span.total-item-nums')
        max_page = 1
        # if len(elements) == 0:
        #     max_page = 0
        #     print('该关键词搜索不到商品')
        # else:
        #     max_page = int(elements[0].text.replace('共', '').replace('页', ''))
        if(end_page>max_page):
            end_page = max_page
    except:
        driver.quit()
        print('与现有浏览器连接断开')

    for page in range(start_page, end_page+1):
        try:
            [single_total_num, single_record_num] = scrape_single_page(driver, keyword, page, file_name, headers)
            total_num += single_total_num
            record_num += single_record_num
        except Exception as e:
            print(e)
            driver.quit()
            print('与现有浏览器连接断开')
            break
        except KeyboardInterrupt:
            driver.quit()
            print('用户主动中断爬虫，与现有浏览器连接断开')
            break

    driver.quit()
    print('与现有浏览器连接断开')
    # 重命名文件
    new_file_name = f"data/wph/唯品会_{urllib.parse.unquote(keyword)}_{time_string}_({record_num} of {total_num}).xlsx"
    try:
        os.rename(file_name, new_file_name)
        print(f"已将文件 {file_name} 重命名为 {new_file_name}")
    except Exception as e:
        print(e)
        print(f"重命名文件 {file_name} 失败")
    return [total_num, record_num]

def scrape_single_page(driver, keyword, page, file_name, headers):
    workbook = load_workbook(file_name)
    sheet = workbook.active
    last_row = sheet.max_row
    total_num = 0
    record_num = 0

    delay = random.uniform(min_delay, max_delay)
    # time.sleep(delay)

    print("正在记录第"+str(page)+"页")
    
    searchUrl = "https://pifa.pinduoduo.com/search?word="+keyword
    driver.get(searchUrl)
    scroll_num = 0
    for i in range(scroll_num):
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(5)
    html = driver.execute_script(
        "return document.documentElement.outerHTML")
    # 创建 Beautiful Soup 对象
    soup = BeautifulSoup(html, "html.parser")

    # 使用 select 方法查找指定的元素
    elements = soup.select('div.goods-item')
    print(len(elements))
    # try:
    #     for (index, element) in enumerate(elements, start=1):
    #         goods_titles = element.select('div.c-goods-item__name')

    #         total_num += 1
    #         # # 筛选
    #         # if (len(goods_titles) != 0):
    #         #     if filter_by_goods_name(goods_titles[0].text):
    #         #         continue

    #         record_num += 1

    #         # 下一行
    #         last_row+=1
    #         last_column = 0

    #         # # 序号
    #         # try:
    #         #     last_column+=1
    #         #     ordinal = last_row-1
    #         #     # sheet.column_dimensions[get_column_letter(last_column)].width = len(str(ordinal)) * 2.5
    #         #     sheet.column_dimensions[get_column_letter(last_column)].width = column_width
    #         #     sheet.row_dimensions[last_row].height = row_height
    #         #     ordinal_cell = sheet[f"{get_column_letter(last_column)}{last_row}"]
    #         #     ordinal_cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
    #         #     sheet.cell(row=last_row, column=last_column, value=ordinal)
    #         # except:
    #         #     print(f'记录“{headers[last_column-1]}”时出错')
    #         #     return

    #         # # 电商平台
    #         # try:
    #         #     last_column+=1
    #         #     platform_name = '唯品会'
    #         #     # sheet.column_dimensions[get_column_letter(last_column)].width = len(platform_name) / 1.5
    #         #     sheet.column_dimensions[get_column_letter(last_column)].width = column_width
    #         #     sheet.row_dimensions[last_row].height = row_height
    #         #     current_time_cell = sheet[f"{get_column_letter(last_column)}{last_row}"]
    #         #     current_time_cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
    #         #     sheet.cell(row=last_row, column=last_column, value=platform_name)
    #         # except:
    #         #     print(f'记录“{headers[last_column-1]}”时出错')
    #         #     return

    #         # # 关键词
    #         # try:
    #         #     last_column+=1
    #         #     search_keyword = urllib.parse.unquote(keyword)
    #         #     # sheet.column_dimensions[get_column_letter(last_column)].width = len(search_keyword) * 2.5
    #         #     sheet.column_dimensions[get_column_letter(last_column)].width = column_width
    #         #     sheet.row_dimensions[last_row].height = row_height
    #         #     search_keyword_cell = sheet[f"{get_column_letter(last_column)}{last_row}"]
    #         #     search_keyword_cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
    #         #     sheet.cell(row=last_row, column=last_column, value=search_keyword)
    #         # except:
    #         #     print(f'记录“{headers[last_column-1]}”时出错')
    #         #     return

    #         # # 店铺名称
    #         # try:
    #         #     last_column+=1
    #         #     sheet.column_dimensions[get_column_letter(last_column)].width = column_width
    #         #     sheet.row_dimensions[last_row].height = row_height
    #         #     manager_cell = sheet[f"{get_column_letter(last_column)}{last_row}"]
    #         #     manager_cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
    #         # except:
    #         #     print(f'记录“{headers[last_column-1]}”时出错')
    #         #     return

    #         # # 店铺网址
    #         # try:
    #         #     last_column+=1
    #         #     sheet.column_dimensions[get_column_letter(last_column)].width = column_width
    #         #     sheet.row_dimensions[last_row].height = row_height
    #         #     manager_cell = sheet[f"{get_column_letter(last_column)}{last_row}"]
    #         #     manager_cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
    #         # except:
    #         #     print(f'记录“{headers[last_column-1]}”时出错')
    #         #     return
            
    #         # # 店铺经营主体信息
    #         # try:
    #         #     last_column+=1
    #         #     sheet.column_dimensions[get_column_letter(last_column)].width = column_width
    #         #     sheet.row_dimensions[last_row].height = row_height
    #         #     manager_cell = sheet[f"{get_column_letter(last_column)}{last_row}"]
    #         #     manager_cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
    #         # except:
    #         #     print(f'记录“{headers[last_column-1]}”时出错')
    #         #     return

    #         # # 商品图片
    #         # try:
    #         #     last_column+=1
    #         #     item = element.select('img.J-goods-item__img')
    #         #     if(len(item) != 0):
    #         #         goods_img_url = item[0].get('data-original')
    #         #         if goods_img_url:
    #         #             goods_img_url = 'https:' + goods_img_url
    #         #             if goods_img_url.endswith('.avif'):
    #         #                 goods_img_url = goods_img_url[:-5]
    #         #         # else:
    #         #         #     goods_img_url = 'https:' + (goods_elements[0].select('img')[0].get('data-lazy-img'))
    #         #         #     if goods_img_url.endswith('.avif'):
    #         #         #         goods_img_url = goods_img_url[:-5]
    #         #         # sheet.column_dimensions[get_column_letter(last_column)].width = len(goods_img_url) / 5
    #         #         sheet.column_dimensions[get_column_letter(last_column)].width = column_width
    #         #         sheet.row_dimensions[last_row].height = row_height
    #         #         goods_img_url_cell = sheet[f"{get_column_letter(last_column)}{last_row}"]
    #         #         goods_img_url_cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
    #         #         goods_img_url_cell.font = Font(underline="single", color="0563C1")
    #         #         goods_img_url_cell.hyperlink = goods_img_url
    #         #         sheet.cell(row=last_row, column=last_column, value=goods_img_url)
    #         # except:
    #         #     print(f'记录“{headers[last_column-1]}”时出错')
    #         #     return

    #         # # 商品标题
    #         # try:
    #         #     last_column+=1
    #         #     if (len(goods_titles) != 0):
    #         #         goods_title = goods_titles[0].text.strip()
    #         #         # sheet.column_dimensions[get_column_letter(last_column)].width = len(goods_title) / 3.5
    #         #         sheet.column_dimensions[get_column_letter(last_column)].width = column_width
    #         #         sheet.row_dimensions[last_row].height = row_height
    #         #         shop_title_cell = sheet[f"{get_column_letter(last_column)}{last_row}"]
    #         #         shop_title_cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
    #         #         sheet.cell(row=last_row, column=last_column, value=goods_title)
    #         #     else:
    #         #         sheet.cell(row=last_row, column=last_column, value='')
    #         # except:
    #         #     print(f'记录“{headers[last_column-1]}”时出错')
    #         #     return

    #         # # 商品品牌
    #         # try:
    #         #     last_column+=1
    #         #     sheet.column_dimensions[get_column_letter(last_column)].width = column_width
    #         #     sheet.row_dimensions[last_row].height = row_height
    #         #     manager_cell = sheet[f"{get_column_letter(last_column)}{last_row}"]
    #         #     manager_cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
    #         # except:
    #         #     print(f'记录“{headers[last_column-1]}”时出错')
    #         #     return

    #         # # 商品链接
    #         # try:
    #         #     last_column+=1
    #         #     item = element.select('a')
    #         #     if (len(item) != 0):
    #         #         goods_link = 'https:'+item[0].get('href')
    #         #         # sheet.column_dimensions[get_column_letter(last_column)].width = len(goods_link) / 2
    #         #         sheet.column_dimensions[get_column_letter(last_column)].width = column_width
    #         #         sheet.row_dimensions[last_row].height = row_height
    #         #         goods_link_cell = sheet[f"{get_column_letter(last_column)}{last_row}"]
    #         #         goods_link_cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
    #         #         goods_link_cell.font = Font(underline="single", color="0563C1")
    #         #         goods_link_cell.hyperlink = goods_link
    #         #         sheet.cell(row=last_row, column=last_column, value=goods_link)
    #         #     else:
    #         #         sheet.cell(row=last_row, column=last_column, value='')
    #         # except:
    #         #     print(f'记录“{headers[last_column-1]}”时出错')
    #         #     return

    #         # # 单价
    #         # try:
    #         #     last_column+=1
    #         #     item = element.select('div.c-goods-item__sale-price.J-goods-item__sale-price')
    #         #     if (len(item) != 0):
    #         #         goods_price = item[0].text
    #         #         goods_price = goods_price.strip().replace('¥', '')
    #         #         # sheet.column_dimensions[get_column_letter(last_column)].width = len(goods_price) * 2
    #         #         sheet.column_dimensions[get_column_letter(last_column)].width = column_width
    #         #         sheet.row_dimensions[last_row].height = row_height
    #         #         goods_price_cell = sheet[f"{get_column_letter(last_column)}{last_row}"]
    #         #         goods_price_cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
    #         #         sheet.cell(row=last_row, column=last_column, value=goods_price)
    #         #     else:
    #         #         sheet.cell(row=last_row, column=last_column, value='')
    #         # except:
    #         #     print(f'记录“{headers[last_column-1]}”时出错')
    #         #     return

    #         # # 销售量
    #         # try:
    #         #     last_column+=1
    #         #     sheet.column_dimensions[get_column_letter(last_column)].width = column_width
    #         #     sheet.row_dimensions[last_row].height = row_height
    #         #     manager_cell = sheet[f"{get_column_letter(last_column)}{last_row}"]
    #         #     manager_cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
    #         # except:
    #         #     print(f'记录“{headers[last_column-1]}”时出错')
    #         #     return
            
    #         # # 商品评论数
    #         # try:
    #         #     last_column+=1
    #         #     sheet.column_dimensions[get_column_letter(last_column)].width = column_width
    #         #     sheet.row_dimensions[last_row].height = row_height
    #         #     manager_cell = sheet[f"{get_column_letter(last_column)}{last_row}"]
    #         #     manager_cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
    #         # except:
    #         #     print(f'记录“{headers[last_column-1]}”时出错')
    #         #     return

    #         # # 销售额
    #         # try:
    #         #     last_column+=1
    #         #     sheet.column_dimensions[get_column_letter(last_column)].width = column_width
    #         #     sheet.row_dimensions[last_row].height = row_height
    #         #     manager_cell = sheet[f"{get_column_letter(last_column)}{last_row}"]
    #         #     manager_cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
    #         # except:
    #         #     print(f'记录“{headers[last_column-1]}”时出错')
    #         #     return


    # except Exception as e:
    #     print('表格记录数据时出错')
    #     driver.quit()
    #     print('与现有浏览器连接断开')


    try:
        workbook.save(file_name)
        print(f"已保存第 {page} 页数据到 {file_name}")
    except Exception as e:
        print('文件保存时出错')
        driver.quit()
        print('与现有浏览器连接断开')
    return [total_num, record_num]

def filter_by_shop_name(shopName):
    keywords = ['华为苏宁易购自营官方旗舰店']
    if shopName in keywords:
        return True
    else:
        return False

def filter_by_goods_name(goodsName):
    required_keywords = 3  # 至少需要满足的关键字数
    lower_case_good_name = goodsName.lower()
    matched_keywords = 0
    keywords = ['xiaomi', 'huawei', 'oppo', 'vivo', 'redmi', 'realme', '真我', '红米', '小米', '华为', '荣耀', '魅族', '一加', '苹果']
    for keyword in keywords:
        lower_case_keyword = keyword.lower()
        if lower_case_keyword in lower_case_good_name:
            matched_keywords += 1
            if matched_keywords >= required_keywords:
                return True
    return False

def filter_by_goods_commit(goodsCommit):
    if not goodsCommit:
        return True
    if goodsCommit.endswith('万+'):
        return False
    elif goodsCommit.endswith('+'):
        if int(goodsCommit[:-1])>=200:
            return False
        else:
            return True
    else:
        return True

def convert_string_to_number(string):
    if not string:
        return 0
    if string.endswith('万+'):
        number = float(string[:-2]) * 10000
    elif string.endswith('+'):
        number = float(string[:-1])
    else:
        number = float(string)
    return number

def is_float(string):
    try:
        float(string)
        return True
    except ValueError:
        return False

if __name__ == "__main__":
    keyword = urllib.parse.quote("华为手表")
    start_page = 1
    end_page = 1
    [total_num, record_num] = scrape_multiple_pages(keyword, start_page, end_page)
    print(f"共找到 {total_num} 条数据，经过筛选，已记录 {record_num} 条数据")
