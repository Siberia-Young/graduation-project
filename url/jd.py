from openpyxl import load_workbook
from selenium import webdriver
from bs4 import BeautifulSoup
import time

# filename = input("请输入文件名称：")
filename = "data/jd/京东_华为_2023-11-05_14-30-39_(1837 of 3000).xlsx"
workbook = load_workbook(filename)
sheet = workbook.active
start_time = time.time()

start_row = 1255
end_row = sheet.max_row
aim_col = 10

total = end_row - start_row + 1
current = 0

options = webdriver.FirefoxOptions()
driver = webdriver.Remote(command_executor="http://127.0.0.1:4444", options=options)

# options = webdriver.FirefoxOptions()
# driver = webdriver.Firefox(options=options)

try:
    for row in range(start_row, end_row + 1):
        current+=1
        print(f"\r当前进度：{current}/{total}", end="")
        goods_link = sheet.cell(row=row, column=aim_col).value
        driver.get(goods_link)
        tempHTML = driver.execute_script("return document.documentElement.outerHTML")
        tempSoup = BeautifulSoup(tempHTML, "html.parser")
        try:
            goods_brand_element = tempSoup.find_all('ul',id='parameter-brand')
            goods_brand = len(goods_brand_element) == 0 and '暂无' or goods_brand_element[0].select('li a')[0].text
            sheet.cell(row=row, column=9, value=goods_brand)
        except:
            workbook.save(filename)
            driver.quit()
            print('与现有浏览器连接断开')
        time.sleep(0.2)
except:
    print('连接超时')
    print('主动中断')
finally:
    # 保存文件
    workbook.save(filename)
    driver.quit()
    print('与现有浏览器连接断开')
    end_time = time.time()
    duration = end_time - start_time
    print(f"爬虫耗时：{duration:.2f} 秒")
    print(f"目标数量：{total} 条")
    print(f"已获取数量：{current} 条")
    unit = current / (duration / 60)
    print(f"每分钟爬取数量：{unit:.2f} 条")