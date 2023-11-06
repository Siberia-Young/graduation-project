from openpyxl import load_workbook
from selenium import webdriver
from bs4 import BeautifulSoup
import time

# filename = input("请输入文件名称：")
filename = "data/wph/唯品会_华为汽车配件_2023-11-05_22-00-49_(80 of 80).xlsx"
workbook = load_workbook(filename)
sheet = workbook.active
start_time = time.time()

start_row = 2
end_row = sheet.max_row
aim_col = 10

total = end_row - start_row + 1
current = 0

# options = webdriver.FirefoxOptions()
# driver = webdriver.Remote(command_executor="http://127.0.0.1:4444", options=options)

options = webdriver.FirefoxOptions()
driver = webdriver.Firefox(options=options)
driver.get('https://passport.vip.com/login')
time.sleep(20)

def convert_string_to_number(string):
    if not string:
        return 0
    if string.endswith('万+'):
        number = float(string[:-2]) * 10000
    elif string.endswith('+'):
        if string[:-1] == '999':
            number = float(string[:-1]) + 1
        else:
            number = float(string[:-1])
    else:
        number = float(string)
    return number

try:
    for row in range(start_row, end_row + 1):
        current+=1
        print(f"\r当前进度：{current}/{total}", end="")
        goods_link = sheet.cell(row=row, column=aim_col).value
        driver.get(goods_link)
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(1)
        tempHTML = driver.execute_script("return document.documentElement.outerHTML")
        tempSoup = BeautifulSoup(tempHTML, "html.parser")
        try:
            goods_brand_element = tempSoup.find_all('div',id='J_detail_info_mation')
            goods_brand_element = goods_brand_element[0].select('a.pib-title-class.J_brandName')
            goods_brand = len(goods_brand_element) == 0 and '暂无' or goods_brand_element[0].text
            sheet.cell(row=row, column=9, value=goods_brand)

            goods_commit_element = tempSoup.select('i.J-detail-commentCnt-count')
            goods_commit = len(goods_commit_element) == 0 and 0 or goods_commit_element[0].text
            sheet.cell(row=row, column=13, value=goods_commit)
            try:
                sheet.cell(row=row, column=14, value=float(sheet.cell(row=row, column=11).value)*convert_string_to_number(goods_commit))
            except:
                print('记录出错')
                break
        except:
            workbook.save(filename)
            driver.quit()
            print('与现有浏览器连接断开')
except:
    print('连接超时')
    print('主动中断')
finally:
    # 保存文件
    workbook.save(filename)
    driver.quit()
    print('与现有浏览器连接断开')
    end_time = time.time()
    duration = end_time - start_time
    print(f"爬虫耗时：{duration:.2f} 秒")
    print(f"目标数量：{total} 条")
    print(f"已获取数量：{current} 条")
    unit = current / (duration / 60)
    print(f"每分钟爬取数量：{unit:.2f} 条")