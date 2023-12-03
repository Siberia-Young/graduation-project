from openpyxl import load_workbook
from openpyxl import Workbook
import time
import os
from openpyxl.utils.cell import get_column_letter
import re

file_name = "data/pdd/merge/841.xlsx"
num = 3
new_file_name = file_name.replace('.xlsx','_') + str(num) + '.xlsx'

# 打开需读取的excel表
workbook = load_workbook(file_name)
sheet = workbook.active

# 新建excel表
new_workbook = Workbook()
new_sheet = new_workbook.active

# 处理表头
print(f'\n正在处理表头')
first_row = sheet[1]
for cell in first_row:
    new_sheet[cell.coordinate].value = cell.value

# 通过店铺白名单筛选
try:
    def convert_string_to_number(string):
        if not string:
            return 0
        if isinstance(string, int):
            return string
        if string.endswith('万+'):
            number = float(string[:-2]) * 10000
        elif string.endswith('万'):
            number = float(string[:-1]) * 10000
        elif string.endswith('+'):
            number = float(string[:-1])
        else:
            number = float(string)
        return number
    def check_keywords1(text):
        keywords = ['手表','watch','gt']
        pattern = '|'.join(keywords)
        match = re.search(pattern, text, flags=re.IGNORECASE)
        return match is not None
    def check_keywords2(text):
        keywords = ['HUAWEI','华为']
        pattern = '|'.join(keywords)
        match = re.search(pattern, text, flags=re.IGNORECASE)
        return match is not None
    list = ['华为官方旗舰店','荣耀官方旗舰店','天天特卖工厂店','天猫超市','绿联数码旗舰店']
    record_list = []
    start_row = 2
    end_row = sheet.max_row

    total = end_row - start_row + 1
    current = 0
    start_time = time.time()
    time.sleep(1)
    print(f'\n正在通过店铺白名单筛选')
    for row in range(start_row, end_row + 1):
        shop_name = sheet.cell(row=row, column=4).value
        goods_title = sheet.cell(row=row, column=8).value
        goods_nums = sheet.cell(row=row, column=12).value
        if convert_string_to_number(goods_nums) >= 200 and check_keywords1(goods_title) and check_keywords2(goods_title):
            record_list.append(row)
except Exception as e:
    print(e)
    print('通过店铺白名单筛选时出错')

# 记录数据到新表
try:
    start_row = 2
    end_row = sheet.max_row

    total = len(record_list)
    current = 0
    start_time = time.time()
    time.sleep(1)
    print(f'\n正在记录数据到新表')
    for row in record_list:
        current+=1
        res = (total - current) / (current / ((time.time() - start_time) / 60))
        print(f"\r当前进度：{current}/{total}，预计仍需：{res:.2f} min", end="")
        for cell in sheet[row]:
            new_sheet[f"{get_column_letter(cell.column)}{current+1}"].value = cell.value
except Exception as e:
    print(e)
    print('记录数据到新表时出错')

# 处理序号
try:
    start_row = 2
    end_row = new_sheet.max_row

    total = total
    current = 0
    start_time = time.time()
    time.sleep(1)
    print(f'\n正在处理序号')
    for row in range(start_row, end_row + 1):
        current+=1
        res = (total - current) / (current / ((time.time() - start_time) / 60))
        print(f"\r当前进度：{current}/{total}，预计仍需：{res:.2f} min", end="")
        new_sheet.cell(row=row, column=1, value=row-1)
except Exception as e:
    print(e)
    print('处理序号时出错')

new_workbook.save(new_file_name)

# 修改文件名
try:
    temp_file_name = "/".join(file_name.split("/")[:-1]) + '/temp.xlsx'
    os.rename(file_name, temp_file_name)
    os.rename(new_file_name, file_name)
    os.rename(temp_file_name, new_file_name)
except Exception as e:
    print(e)
    print('修改文件名时出错')