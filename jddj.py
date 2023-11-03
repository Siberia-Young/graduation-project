from seleniumwire import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver import ActionChains
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as ExcelImage
from io import BytesIO
from PIL import Image as PILImage
import os
import io
import urllib.request
import urllib.parse
import time
import random
import datetime
import gzip
import json
import demjson

ua = 'Mozilla/5.0 (iPhone; CPU iPhone OS 15_0 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Edg/94.0.992.31 Mobile/15E148 Safari/604.1'


def open_browser(keyword, start_page, end_page):
    # 创建带有Selenium Wire的Edge WebDriver对象
    options = webdriver.EdgeOptions()
    options.add_argument("--user-agent={}".format(ua))  # 设置用户代理为iPhone的User Agent
    driver = webdriver.Edge(options=options)
    driver.set_window_size(360, 820)
    # 执行浏览器操作
    time.sleep(2)
    print("打开浏览器")
    driver.get('https://daojia.jd.com')
    time.sleep(5)
    driver.get('https://daojia.jd.com/html/index/IndexSearch?keyword={}'.format(keyword))
    time.sleep(5)
    return scrape_multiple_pages(driver, keyword, start_page, end_page)

def scrape_multiple_pages(driver, keyword, start_page, end_page):
    # 获取当前时间
    current_time = datetime.datetime.now()
    # 格式化时间字符串
    time_string = current_time.strftime("%Y-%m-%d_%H-%M-%S")
    # 构建文件名
    file_name = f"data/jddj/京东到家_{urllib.parse.unquote(keyword)}_{time_string}.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    headers = ['序号', '平台', '店铺名称', '店铺链接', '商品ID', '商品标题', '商品链接', '商品首页图', '现商品单价', '商品目前销量']
    sheet.append(headers)
    workbook.save(file_name)

    for page in range(start_page, end_page+1):
        try:
            scrape_single_page(driver, page)
        except KeyboardInterrupt:
            driver.quit()
            print('用户主动中断爬虫')
            break

    [total_num, record_num] = record_data(driver, file_name, end_page - start_page + 1)
    driver.quit()
    print('与现有浏览器连接断开')
    # 重命名文件
    new_file_name = f"data/jddj/京东到家_{urllib.parse.unquote(keyword)}_{time_string}_({record_num} of {total_num}).xlsx"
    try:
        os.rename(file_name, new_file_name)
        print(f"已将文件 {file_name} 重命名为 {new_file_name}")
    except Exception as e:
        print(e)
        print(f"重命名文件 {file_name} 失败")
    return [total_num, record_num]


def scrape_single_page(driver, page):
    # 模拟滑轮滚动
    print('正在获取第 {} 页数据'.format(page))
    element = driver.find_element(By.CSS_SELECTOR, '.css-1dbjc4n.r-150rngu.r-eqz5dr.r-16y2uox.r-1wbh5a2.r-11yh6sk.r-1rnoaur.r-2eszeu.r-1sncvnh')
    script = "arguments[0].scrollTop = arguments[0].scrollHeight;"
    driver.execute_script(script, element)
    time.sleep(5)


def record_data(driver, file_name, total_page):
    logs = driver.requests
    logs_list = []
    total_num = 0
    record_num = 0
    for log in logs:
        if 'https://api.m.jd.com/client.action?appid=JDReactDaoJiaH5&functionId=dj_homeSearch_searchSkuResultByTab' in log.url and log.method == 'POST':
            logs_list.append(json.loads(demjson.encode(demjson.decode(gzip.decompress(log.response.body).decode('utf-8')))))
    for i in range(total_page):
        if i < len(logs_list):
            workbook = load_workbook(file_name)
            sheet = workbook.active
            last_row = sheet.max_row
            print(f'正在记录第 {i+1} 页，本页未经筛选的数据为 {len(logs_list[i]["result"]["searchResultVOList"])} 条')
            for index,item in enumerate(logs_list[i]['result']['searchResultVOList'], start=1):
                total_num+=1
                # 筛选
                if ('storeName' in item.keys()):
                    if filter_by_shop_name(item['storeName']):
                        continue
                if ('skuName' in item.keys()):
                    if filter_by_goods_name(item['skuName']):
                        continue
                record_num+=1

                # 下一行
                last_row+=1
                last_column = 0
                # 序号
                last_column+=1
                ordinal = last_row-1
                sheet.column_dimensions[get_column_letter(last_column)].width = len(str(ordinal)) * 2.5
                ordinal_cell = sheet[f"{get_column_letter(last_column)}{last_row}"]
                ordinal_cell.alignment = Alignment(wrapText=True, vertical='center')
                sheet.cell(row=last_row, column=last_column, value=ordinal)

                # 平台名称
                last_column+=1
                platform_name = '京东到家'
                sheet.column_dimensions[get_column_letter(last_column)].width = len(platform_name) * 3
                platform_name_cell = sheet[f"{get_column_letter(last_column)}{last_row}"]
                platform_name_cell.alignment = Alignment(wrapText=True, vertical='center')
                sheet.cell(row=last_row, column=last_column, value=platform_name)
  
                # 店铺名称
                last_column+=1
                shop_name = 'storeName' in item.keys() and item['storeName'] or '暂无店铺名称'
                sheet.column_dimensions[get_column_letter(last_column)].width = len(shop_name) * 2
                shop_name_cell = sheet[f"{get_column_letter(last_column)}{last_row}"]
                shop_name_cell.alignment = Alignment(wrapText=True, vertical='center')
                sheet.cell(row=last_row, column=last_column, value=shop_name)

                # 店铺链接
                last_column+=1
                shop_link = 'storeId' in item.keys() and ('https://daojia.jd.com/html/index/storeHome?storeId='+item['storeId']) or '暂无店铺链接'
                sheet.column_dimensions[get_column_letter(last_column)].width = len(shop_link) / 2
                shop_link_cell = sheet[f"{get_column_letter(last_column)}{last_row}"]
                shop_link_cell.alignment = Alignment(wrapText=True, vertical='center')
                shop_link_cell.font = Font(underline="single", color="0563C1")
                shop_link_cell.hyperlink = shop_link
                sheet.cell(row=last_row, column=last_column, value=shop_link)
                
                # 商品ID
                last_column+=1
                goods_id = 'skuId' in item.keys() and item['skuId'] or '暂无商品ID'
                sheet.column_dimensions[get_column_letter(last_column)].width = len(goods_id) * 1.5
                goods_id_cell = sheet[f"{get_column_letter(last_column)}{last_row}"]
                goods_id_cell.alignment = Alignment(wrapText=True, vertical='center')
                sheet.cell(row=last_row, column=last_column, value=goods_id)

                # 商品名称
                last_column+=1
                shop_name = 'skuName' in item.keys() and item['skuName'] or '暂无商品名称'
                sheet.column_dimensions[get_column_letter(last_column)].width = len(shop_name) / 1.2
                shop_name_cell = sheet[f"{get_column_letter(last_column)}{last_row}"]
                shop_name_cell.alignment = Alignment(wrapText=True, vertical='center')
                sheet.cell(row=last_row, column=last_column, value=shop_name)

                # 商品链接
                last_column+=1
                goods_link = 'storeId' in item.keys() and 'skuId' in item.keys() and 'spuId' in item.keys() and ('https://daojia.jd.com/html/index/goodsDetails?storeId='+item['storeId']+'&skuId='+item['skuId']+'&spuId='+str(item['spuId'])) or '暂无商品链接'
                sheet.column_dimensions[get_column_letter(last_column)].width = len(goods_link) / 6
                goods_link_cell = sheet[f"{get_column_letter(last_column)}{last_row}"]
                goods_link_cell.alignment = Alignment(wrapText=True, vertical='center')
                goods_link_cell.font = Font(underline="single", color="0563C1")
                goods_link_cell.hyperlink = goods_link
                sheet.cell(row=last_row, column=last_column, value=goods_link)

                # 商品首页图
                last_column+=1
                if 'spuImg' in item.keys():
                    headers = {
                        'User-Agent': ua
                    }
                    request = urllib.request.Request(url=item['spuImg'],headers=headers)
                    response = urllib.request.urlopen(request)
                    content = response.read()
                    image = PILImage.open(BytesIO(content))
                    # img_width, img_height = image.size
                    # new_width = img_width // 2
                    # new_height = img_height // 2
                    # resized_image = image.resize((new_width, new_height))
                    goods_img = ExcelImage(image)
                    goods_img_cell = sheet.cell(row=last_row, column=last_column)
                    sheet[f"{get_column_letter(last_column)}{last_row}"].alignment = Alignment(vertical='center')
                    sheet.add_image(goods_img, goods_img_cell.coordinate)
                    sheet.column_dimensions[goods_img_cell.column_letter].width = goods_img.width / 7.2
                    sheet.row_dimensions[goods_img_cell.row].height = goods_img.height / 1.32
                else:
                    sheet.cell(row=last_row, column=last_column, value='暂无图片')

                # 商品售价
                last_column+=1
                goods_price = 'realTimePrice' in item.keys() and item['realTimePrice'] or '暂无报价'
                sheet.column_dimensions[get_column_letter(last_column)].width = len(goods_id) * 1.0
                goods_price_cell = sheet[f"{get_column_letter(last_column)}{last_row}"]
                goods_price_cell.alignment = Alignment(wrapText=True, vertical='center')
                sheet.cell(row=last_row, column=last_column, value=goods_price)          

                # 商品销量
                last_column+=1
                goods_sale = 'monthSales' in item.keys() and item['monthSales'] or '暂无销量'
                sheet.column_dimensions[get_column_letter(last_column)].width = len(goods_id) * 1.5
                goods_sale_cell = sheet[f"{get_column_letter(last_column)}{last_row}"]
                goods_sale_cell.alignment = Alignment(wrapText=True, vertical='center')
                sheet.cell(row=last_row, column=last_column, value=goods_sale)

            workbook.save(file_name)
            print(f"已保存第 {i+1} 页数据到 {file_name}")
    return [total_num, record_num]

def filter_by_shop_name(shopName):
    keywords = ['华为京东自营官方旗舰店']
    if shopName in keywords:
        return True
    else:
        return False

def filter_by_goods_name(goodsName):
    required_keywords = 3  # 至少需要满足的关键字数
    lower_case_good_name = goodsName.lower()
    matched_keywords = 0
    keywords = ['xiaomi', 'huawei', 'oppo', 'vivo', 'redmi', 'realme', '真我', '红米', '小米', '华为', '荣耀', '魅族', '一加', '苹果']
    for keyword in keywords:
        lower_case_keyword = keyword.lower()
        if lower_case_keyword in lower_case_good_name:
            matched_keywords += 1
            if matched_keywords >= required_keywords:
                return True
    return False

if __name__ == "__main__":
    keyword = urllib.parse.quote("华为适用")
    start_page = 1
    end_page = 1
    [total_num, record_num] = open_browser(keyword, start_page, end_page)
    print(f"共找到 {total_num} 条数据，经过筛选，已保存 {record_num} 条数据")
