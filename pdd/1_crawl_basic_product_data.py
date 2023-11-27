from openpyxl import load_workbook
from selenium import webdriver
from openpyxl import Workbook
from bs4 import BeautifulSoup
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import json
import os
import requests
import urllib.parse
import time
import random
import datetime
import re

row_height = 40
column_width = 14

# 读取JSON文件
list = []
with open('data/pdd/json/5.json', 'r', encoding='utf-8') as file:
    list = json.load(file)

keyword = input('请输入关键词：')
current_time = datetime.datetime.now()
time_string = current_time.strftime("%Y-%m-%d_%H-%M-%S")
file_name = f"data/pdd/拼多多_{keyword}_{time_string}.xlsx"
total_num = 0
record_num = 0

workbook = Workbook()
sheet = workbook.active
headers = ['序号', '电商平台', '关键词', '店铺名称', '店铺网址', '店铺经营主体信息', '商品图片', '商品标题', '商品品牌', '商品链接', '单价', '销售量', '商品评论数', '销售额']

sheet.append(headers)
for index, cell in enumerate(sheet[1], start=1):
    if index == 16:
        cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        cell.font = Font(bold=True, color="FFFFFF")
    else:
        cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        cell.font = Font(bold=True, color="000000")
    cell.alignment = Alignment(horizontal='center', vertical='center')

workbook.save(file_name)
total_num = 0
record_num = 0
workbook = load_workbook(file_name)
sheet = workbook.active
last_row = sheet.max_row

def convert_string_to_number(string):
    if not string:
        return 0
    if string.endswith('万+'):
        number = float(string[:-2]) * 10000
    elif string.endswith('万'):
        number = float(string[:-1]) * 10000
    elif string.endswith('+'):
        number = float(string[:-1])
    else:
        number = float(string)
    return number

for i in range(len(list)):
    for j in range(len(list[i])):
        total_num += 1
        record_num += 1

        # 下一行
        last_row+=1
        last_column = 0

        # 序号
        try:
            last_column+=1
            ordinal = last_row-1
            # sheet.column_dimensions[get_column_letter(last_column)].width = len(str(ordinal)) * 2.5
            sheet.column_dimensions[get_column_letter(last_column)].width = column_width
            sheet.row_dimensions[last_row].height = row_height
            ordinal_cell = sheet[f"{get_column_letter(last_column)}{last_row}"]
            ordinal_cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
            sheet.cell(row=last_row, column=last_column, value=ordinal)
        except:
            print(f'记录“{headers[last_column-1]}”时出错')
            

        # 电商平台
        try:
            last_column+=1
            platform_name = '拼多多-批发'
            # sheet.column_dimensions[get_column_letter(last_column)].width = len(platform_name) / 1.5
            sheet.column_dimensions[get_column_letter(last_column)].width = column_width
            sheet.row_dimensions[last_row].height = row_height
            current_time_cell = sheet[f"{get_column_letter(last_column)}{last_row}"]
            current_time_cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
            sheet.cell(row=last_row, column=last_column, value=platform_name)
        except:
            print(f'记录“{headers[last_column-1]}”时出错')
            

        # 关键词
        try:
            last_column+=1
            search_keyword = keyword
            # sheet.column_dimensions[get_column_letter(last_column)].width = len(search_keyword) * 2.5
            sheet.column_dimensions[get_column_letter(last_column)].width = column_width
            sheet.row_dimensions[last_row].height = row_height
            search_keyword_cell = sheet[f"{get_column_letter(last_column)}{last_row}"]
            search_keyword_cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
            sheet.cell(row=last_row, column=last_column, value=search_keyword)
        except:
            print(f'记录“{headers[last_column-1]}”时出错')
            

        # 店铺名称
        try:
            last_column+=1
            shop_name = 'mallName' in list[i][j].keys() and list[i][j]['mallName'] or '暂无店铺名称'
            # sheet.column_dimensions[get_column_letter(last_column)].width = len(shop_name) * 1.5
            sheet.column_dimensions[get_column_letter(last_column)].width = column_width
            sheet.row_dimensions[last_row].height = row_height
            shop_name_cell = sheet[f"{get_column_letter(last_column)}{last_row}"]
            shop_name_cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
            sheet.cell(row=last_row, column=last_column, value=shop_name)
        except:
            print(f'记录“{headers[last_column-1]}”时出错')
            

        # 店铺网址
        try:
            last_column+=1
            shop_link = 'mallIdEncrypt' in list[i][j].keys() and ('https://pifa.pinduoduo.com/mall?mid='+list[i][j]['mallIdEncrypt']) or '暂无店铺链接'
            # sheet.column_dimensions[get_column_letter(last_column)].width = len(shop_link) / 3
            sheet.column_dimensions[get_column_letter(last_column)].width = column_width
            sheet.row_dimensions[last_row].height = row_height
            shop_link_cell = sheet[f"{get_column_letter(last_column)}{last_row}"]
            shop_link_cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
            shop_link_cell.font = Font(underline="single", color="0563C1")
            shop_link_cell.hyperlink = shop_link
            sheet.cell(row=last_row, column=last_column, value=shop_link)
        except:
            print(f'记录“{headers[last_column-1]}”时出错')
            
        
        # 店铺经营主体信息
        try:
            last_column+=1
            sheet.column_dimensions[get_column_letter(last_column)].width = column_width
            sheet.row_dimensions[last_row].height = row_height
            manager_cell = sheet[f"{get_column_letter(last_column)}{last_row}"]
            manager_cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
        except:
            print(f'记录“{headers[last_column-1]}”时出错')
            

        # 商品图片
        try:
            last_column+=1
            goods_img_url = 'goodsImgUrl' in list[i][j].keys() and list[i][j]['goodsImgUrl'] or '暂无商品图片'
            sheet.column_dimensions[get_column_letter(last_column)].width = column_width
            sheet.row_dimensions[last_row].height = row_height
            goods_img_url_cell = sheet[f"{get_column_letter(last_column)}{last_row}"]
            goods_img_url_cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
            goods_img_url_cell.font = Font(underline="single", color="0563C1")
            goods_img_url_cell.hyperlink = goods_img_url
            sheet.cell(row=last_row, column=last_column, value=goods_img_url)
        except:
            print(f'记录“{headers[last_column-1]}”时出错')

        # 商品标题
        try:
            last_column+=1
            goods_title = 'goodsName' in list[i][j].keys() and list[i][j]['goodsName'] or '暂无商品标题'
            # sheet.column_dimensions[get_column_letter(last_column)].width = len(goods_title) / 3.5
            sheet.column_dimensions[get_column_letter(last_column)].width = column_width
            sheet.row_dimensions[last_row].height = row_height
            shop_title_cell = sheet[f"{get_column_letter(last_column)}{last_row}"]
            shop_title_cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
            sheet.cell(row=last_row, column=last_column, value=goods_title)
        except:
            print(f'记录“{headers[last_column-1]}”时出错')
            

        # 商品品牌
        try:
            last_column+=1
            sheet.column_dimensions[get_column_letter(last_column)].width = column_width
            sheet.row_dimensions[last_row].height = row_height
            goods_brand = sheet[f"{get_column_letter(last_column)}{last_row}"]
            goods_brand.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
        except:
            print(f'记录“{headers[last_column-1]}”时出错')
            

        # 商品链接
        try:
            last_column+=1
            goods_link = 'goodsId' in list[i][j].keys() and 'https://pifa.pinduoduo.com/goods/detail/?gid='+str(list[i][j]['goodsId']) or '暂无商品链接'
            # sheet.column_dimensions[get_column_letter(last_column)].width = len(goods_link) / 2
            sheet.column_dimensions[get_column_letter(last_column)].width = column_width
            sheet.row_dimensions[last_row].height = row_height
            goods_link_cell = sheet[f"{get_column_letter(last_column)}{last_row}"]
            goods_link_cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
            goods_link_cell.font = Font(underline="single", color="0563C1")
            goods_link_cell.hyperlink = goods_link
            sheet.cell(row=last_row, column=last_column, value=goods_link)
        except:
            print(f'记录“{headers[last_column-1]}”时出错')
            

        # 单价
        try:
            last_column+=1
            goods_price = 'goodsWholeSalePrice' in list[i][j].keys() and (list[i][j]['goodsWholeSalePrice']/100) or '暂无单价'
            # sheet.column_dimensions[get_column_letter(last_column)].width = len(goods_price) * 2
            sheet.column_dimensions[get_column_letter(last_column)].width = column_width
            sheet.row_dimensions[last_row].height = row_height
            goods_price_cell = sheet[f"{get_column_letter(last_column)}{last_row}"]
            goods_price_cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
            sheet.cell(row=last_row, column=last_column, value=goods_price)
        except:
            print(f'记录“{headers[last_column-1]}”时出错')
            

        # 销售量
        try:
            last_column+=1
            goods_num = 'salesTipAmount' in list[i][j].keys() and list[i][j]['salesTipAmount'] or '0'
            # sheet.column_dimensions[get_column_letter(last_column)].width = len(goods_commit) * 2
            sheet.column_dimensions[get_column_letter(last_column)].width = column_width
            sheet.row_dimensions[last_row].height = row_height
            goods_num_cell = sheet[f"{get_column_letter(last_column)}{last_row}"]
            goods_num_cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
            sheet.cell(row=last_row, column=last_column, value=goods_num)
        except:
            print(f'记录“{headers[last_column-1]}”时出错')
            
        
        # 商品评论数
        try:
            last_column+=1
            sheet.column_dimensions[get_column_letter(last_column)].width = column_width
            sheet.row_dimensions[last_row].height = row_height
            goods_commit_cell = sheet[f"{get_column_letter(last_column)}{last_row}"]
            goods_commit_cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
        except:
            print(f'记录“{headers[last_column-1]}”时出错')
            
        # 销售额
        try:
            last_column+=1
            goods_sales = goods_price * convert_string_to_number(goods_num)
            # sheet.column_dimensions[get_column_letter(last_column)].width = len(str(goods_sales)) * 1.5
            sheet.column_dimensions[get_column_letter(last_column)].width = column_width
            sheet.row_dimensions[last_row].height = row_height
            goods_sales_cell = sheet[f"{get_column_letter(last_column)}{last_row}"]
            goods_sales_cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
            sheet.cell(row=last_row, column=last_column, value=goods_sales)
        except:
            print(f'记录“{headers[last_column-1]}”时出错')
                


workbook.save(file_name)