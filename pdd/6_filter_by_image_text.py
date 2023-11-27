from openpyxl import load_workbook
from openpyxl import Workbook
import time
import os
from openpyxl.utils.cell import get_column_letter
import re

file_name = "data/pdd/merge/pdd.xlsx"
num = 6
new_file_name = file_name.replace('.xlsx','_') + str(num) + '.xlsx'

# 打开需读取的excel表
workbook = load_workbook(file_name)
sheet = workbook.active

# 新建excel表
new_workbook = Workbook()
new_sheet = new_workbook.active

# 处理表头
print(f'\n正在处理表头')
first_row = sheet[1]
for cell in first_row:
    new_sheet[cell.coordinate].value = cell.value

# 筛选出符合条件的行
try:
    list = []
    def check_keywords(text):
        keywords = ['适用','官方','1：1','1:1','正品','复刻','huawei','华为']
        pattern = '|'.join(keywords)
        match = re.search(pattern, text, flags=re.IGNORECASE)
        return match is not None
    start_row = 2
    end_row = sheet.max_row

    total = end_row - start_row + 1
    current = 0
    start_time = time.time()
    time.sleep(1)
    print(f'\n正在筛选出符合条件的行')
    for row in range(start_row, end_row + 1):
        current+=1
        res = (total - current) / (current / ((time.time() - start_time) / 60))
        print(f"\r当前进度：{current}/{total}，预计仍需：{res:.2f} min", end="")
        sentence = sheet.cell(row=row, column=16).value
        if sentence is None or (check_keywords(sentence)):
            list.append(row)
except Exception as e:
    print(e)
    print('筛选出符合条件的行时出错')

# 通过图片文字筛选
try:
    start_row = 2
    end_row = sheet.max_row

    total = len(list)
    current = 0
    start_time = time.time()
    time.sleep(1)
    print(f'\n正在通过图片文字筛选')
    for row in list:
        current+=1
        res = (total - current) / (current / ((time.time() - start_time) / 60))
        print(f"\r当前进度：{current}/{total}，预计仍需：{res:.2f} min", end="")
        for cell in sheet[row][:-2]:
            new_sheet[f"{get_column_letter(cell.column)}{current+1}"].value = cell.value
except Exception as e:
    print(e)
    print('通过图片文字筛选时出错')

# 处理序号
try:
    start_row = 2
    end_row = new_sheet.max_row

    total = total
    current = 0
    start_time = time.time()
    time.sleep(1)
    print(f'\n正在处理序号')
    for row in range(start_row, end_row + 1):
        current+=1
        res = (total - current) / (current / ((time.time() - start_time) / 60))
        print(f"\r当前进度：{current}/{total}，预计仍需：{res:.2f} min", end="")
        new_sheet.cell(row=row, column=1, value=row-1)
except Exception as e:
    print(e)
    print('处理序号时出错')

new_workbook.save(new_file_name)

# # 修改文件名
# try:
#     temp_file_name = "/".join(file_name.split("/")[:-1]) + '/temp.xlsx'
#     os.rename(file_name, temp_file_name)
#     os.rename(new_file_name, file_name)
#     os.rename(temp_file_name, new_file_name)
# except Exception as e:
#     print(e)
#     print('修改文件名时出错')