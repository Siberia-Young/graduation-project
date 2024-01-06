from openpyxl import load_workbook
from openpyxl import Workbook
import time
import os
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles import PatternFill

row_height = 40
column_width = 14

num = 2
new_file_name = 'data/1688/merge/merge.xlsx'
folder_path = 'data/1688/merge'

def merge(new_file_name, folder_path):
    # 新建excel表
    new_workbook = Workbook()
    new_sheet = new_workbook.active

    file_num = 0
    for filename in os.listdir(folder_path):
        if filename.endswith(".xlsx"):
            file_num += 1
    
    # 处理表头
    headers = ['序号', '电商平台', '关键词/产品', '店铺名称(全称)', '店铺网址', '店铺经营主体信息', '商品图片', '商品标题', '实际品牌', '商品链接', '价格(单位：元)', '销售量(单位：件)', '商品评价(单位：个)', '销售额(单位：元)']
    new_sheet.append(headers)
    new_workbook.save(new_file_name)
    
    # 记录到新表
    try:
        for index in range(1, file_num+1):
            file_name = os.path.join(folder_path,f'1688 ({index}).xlsx')
            # 打开需读取的excel表
            workbook = load_workbook(file_name)
            sheet = workbook.active
            new_workbook = load_workbook(new_file_name)
            new_sheet = new_workbook.active

            start_row = 2
            end_row = sheet.max_row
            last_row = new_sheet.max_row

            total = end_row - start_row + 1
            current = 0
            start_time = time.time()
            time.sleep(1)
            print(f'\n正在记录数据到新表')
            for row in range(start_row, end_row + 1):
                current+=1
                res = (total - current) / (current / ((time.time() - start_time) / 60))
                print(f"\r当前进度：{current}/{total}，预计仍需：{res:.2f} min", end="")
                for cell in sheet[row]:
                    new_sheet[f"{get_column_letter(cell.column)}{last_row+current}"].value = cell.value
            new_workbook.save(new_file_name)
    except Exception as e:
        print(e)
        print('记录到新表时出错')

    # 处理序号
    try:
        start_row = 2
        end_row = new_sheet.max_row

        total = end_row - start_row + 1
        current = 0
        start_time = time.time()
        time.sleep(1)
        print(f'\n正在处理序号')
        for row in range(start_row, end_row + 1):
            current+=1
            res = (total - current) / (current / ((time.time() - start_time) / 60))
            print(f"\r当前进度：{current}/{total}，预计仍需：{res:.2f} min", end="")
            new_sheet.cell(row=row, column=1, value=row-1)
    except Exception as e:
        print(e)
        print('处理序号时出错')
    finally:
        new_workbook.save(new_file_name)

    # # 处理表头
    # print(f'\n正在处理表头')
    # for index, cell in enumerate(new_sheet[1], start=1):
    #     cell.fill = PatternFill(start_color='EB2B2B', end_color='EB2B2B', fill_type='solid')
    #     cell.font = Font(bold=True, color="FFFFFF")
    #     cell.alignment = Alignment(horizontal='center', vertical='center')

    # # 更新单元格样式
    # try:
    #     list = []
    #     start_row = 2
    #     end_row = new_sheet.max_row

    #     total = end_row - start_row + 1
    #     current = 0
    #     start_time = time.time()
    #     time.sleep(1)
    #     print(f'\n正在更新单元格样式')
    #     for row in range(start_row, end_row + 1):
    #         current+=1
    #         res = (total - current) / (current / ((time.time() - start_time) / 60))
    #         print(f"\r当前进度：{current}/{total}，预计仍需：{res:.2f} min", end="")

    #         # 序号
    #         new_sheet.column_dimensions[get_column_letter(1)].width = column_width
    #         new_sheet.row_dimensions[row].height = row_height
    #         ordinal_cell = new_sheet[f"{get_column_letter(1)}{row}"]
    #         ordinal_cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
    #         new_sheet.cell(row=row, column=1, value=row-1)

    #         # 电商平台
    #         new_sheet.column_dimensions[get_column_letter(2)].width = column_width
    #         new_sheet.row_dimensions[row].height = row_height
    #         platform_name_cell = new_sheet[f"{get_column_letter(2)}{row}"]
    #         platform_name_cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')

    #         # 关键词
    #         new_sheet.column_dimensions[get_column_letter(3)].width = column_width
    #         new_sheet.row_dimensions[row].height = row_height
    #         search_keyword_cell = new_sheet[f"{get_column_letter(3)}{row}"]
    #         search_keyword_cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')

    #         # 店铺名称
    #         new_sheet.column_dimensions[get_column_letter(4)].width = column_width
    #         new_sheet.row_dimensions[row].height = row_height
    #         shop_name_cell = new_sheet[f"{get_column_letter(4)}{row}"]
    #         shop_name_cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')

    #         # 店铺网址*
    #         shop_link_url = new_sheet.cell(row=row, column=5).value
    #         new_sheet.column_dimensions[get_column_letter(5)].width = column_width
    #         new_sheet.row_dimensions[row].height = row_height
    #         shop_link_url_cell = new_sheet[f"{get_column_letter(5)}{row}"]
    #         shop_link_url_cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
    #         shop_link_url_cell.font = Font(underline="single", color="0563C1")
    #         shop_link_url_cell.hyperlink = shop_link_url
    #         new_sheet.cell(row=row, column=5, value=shop_link_url)

    #         # 店铺经营主体信息
    #         new_sheet.column_dimensions[get_column_letter(6)].width = column_width
    #         new_sheet.row_dimensions[row].height = row_height
    #         manager_cell = new_sheet[f"{get_column_letter(6)}{row}"]
    #         manager_cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')

    #         # 商品图片*
    #         goods_img_url = new_sheet.cell(row=row, column=7).value
    #         new_sheet.column_dimensions[get_column_letter(7)].width = column_width
    #         new_sheet.row_dimensions[row].height = row_height
    #         goods_img_url_cell = new_sheet[f"{get_column_letter(7)}{row}"]
    #         goods_img_url_cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
    #         goods_img_url_cell.font = Font(underline="single", color="0563C1")
    #         goods_img_url_cell.hyperlink = goods_img_url
    #         new_sheet.cell(row=row, column=7, value=goods_img_url)

    #         # 商品标题
    #         new_sheet.column_dimensions[get_column_letter(8)].width = column_width
    #         new_sheet.row_dimensions[row].height = row_height
    #         shop_title_cell = new_sheet[f"{get_column_letter(8)}{row}"]
    #         shop_title_cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')

    #         # 商品品牌          
    #         new_sheet.column_dimensions[get_column_letter(9)].width = column_width
    #         new_sheet.row_dimensions[row].height = row_height
    #         goods_brand_cell = new_sheet[f"{get_column_letter(9)}{row}"]
    #         goods_brand_cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')

    #         # 商品链接*
    #         goods_link_url = new_sheet.cell(row=row, column=10).value
    #         new_sheet.column_dimensions[get_column_letter(10)].width = column_width
    #         new_sheet.row_dimensions[row].height = row_height
    #         goods_link_url_cell = new_sheet[f"{get_column_letter(10)}{row}"]
    #         goods_link_url_cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
    #         goods_link_url_cell.font = Font(underline="single", color="0563C1")
    #         goods_link_url_cell.hyperlink = goods_link_url
    #         new_sheet.cell(row=row, column=10, value=goods_link_url)

    #         # 单价
    #         new_sheet.column_dimensions[get_column_letter(11)].width = column_width
    #         new_sheet.row_dimensions[row].height = row_height
    #         goods_price_cell = new_sheet[f"{get_column_letter(11)}{row}"]
    #         goods_price_cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')

    #         # 销售量
    #         new_sheet.column_dimensions[get_column_letter(12)].width = column_width
    #         new_sheet.row_dimensions[row].height = row_height
    #         goods_num_cell = new_sheet[f"{get_column_letter(12)}{row}"]
    #         goods_num_cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
                        
    #         # 商品评论数
    #         new_sheet.column_dimensions[get_column_letter(13)].width = column_width
    #         new_sheet.row_dimensions[row].height = row_height
    #         goods_commit_cell = new_sheet[f"{get_column_letter(13)}{row}"]
    #         goods_commit_cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')

    #         # 销售额
    #         new_sheet.column_dimensions[get_column_letter(14)].width = column_width
    #         new_sheet.row_dimensions[row].height = row_height
    #         goods_sales_cell = new_sheet[f"{get_column_letter(14)}{row}"]
    #         goods_sales_cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
    # except Exception as e:
    #     print(e)
    #     print('更新单元格样式时出错')
    # finally:
    #     # 保存文件
    #     new_workbook.save(new_file_name)