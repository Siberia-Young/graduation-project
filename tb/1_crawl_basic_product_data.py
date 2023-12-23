from selenium import webdriver
from openpyxl import Workbook
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from io import BytesIO
import os
import requests
import urllib.parse
import time
import random
import datetime
import re

min_delay = 5  # 最小延迟时间（单位：秒）
max_delay = 8  # 最大延迟时间（单位：秒）

row_height = 40
column_width = 14

def login(driver):
    print('登录')
    driver.get('https://login.taobao.com/')
    time.sleep(50)

def scrape_multiple_pages(driver, keyword, start_page, end_page):
    # 获取当前时间
    current_time = datetime.datetime.now()
    # 格式化时间字符串
    time_string = current_time.strftime("%Y-%m-%d_%H-%M-%S")
    # 构建文件名
    file_name = f"data/tb/淘宝_{urllib.parse.unquote(keyword)}_{time_string}.xlsx"
    total_num = 0
    record_num = 0

    workbook = Workbook()
    sheet = workbook.active
    headers = ['序号', '电商平台', '关键词/产品', '店铺名称(全称)', '店铺网址', '店铺经营主体信息', '商品图片', '商品标题', '实际品牌', '商品链接', '价格(单位：元)', '销售量(单位：件)', '商品评价(单位：个)', '销售额(单位：元)']

    sheet.append(headers)
    workbook.save(file_name)

    searchInput = driver.find_element("xpath","//input[@id='q']")
    driver.execute_script("arguments[0].setAttribute('autocomplete', 'off')", searchInput)
    driver.execute_script("arguments[0].value = '';", searchInput)
    searchInput.send_keys(urllib.parse.unquote(keyword))

    searchButton = driver.find_element("xpath","//button[@id='button']")
    searchButton.click()

    time.sleep(2)

    li_elements = driver.find_elements(by="xpath", value="//li[contains(@class, 'SortBar--customTabItem--YnxmQgr')]")
    li_elements[1].click()
    time.sleep(3)

    html = driver.execute_script(
        "return document.documentElement.outerHTML")
    # 创建 Beautiful Soup 对象
    soup = BeautifulSoup(html, "html.parser")
    try:
        elements = soup.select('button.next-btn.next-medium.next-btn-normal.next-pagination-item')

        max_page = int(re.search(r"共(\d+)页",elements[1].get('aria-label')).group(1))
        print("最大页数：",max_page,end_page)
        if(end_page>max_page):
            end_page = max_page
    except:
        print('获取最大页数时出错')

    for page in range(start_page, end_page+1):
        try:
            tempHtml = driver.execute_script(
                "return document.documentElement.outerHTML")
            # 创建 Beautiful Soup 对象
            tempSoup = BeautifulSoup(tempHtml, "html.parser")
            try:
                tempElements = tempSoup.select('button.next-btn.next-medium.next-btn-normal.next-pagination-item')
                if len(tempElements) != 0:
                    max_page = int(re.search(r"共(\d+)页",tempElements[1].get('aria-label')).group(1))
                else:
                    max_page = 1
                print('最大页面数：',max_page)
                if(page>max_page):
                    break
            except:
                print('更新最大页数时出错')
            [single_total_num, single_record_num] = scrape_single_page(driver, keyword, start_page, page, file_name, headers)
            total_num += single_total_num
            record_num += single_record_num
            if single_record_num == 0:
                break
        except Exception as e:
            print(e)
            driver.quit()
            print('与现有浏览器连接断开')
            break
        except KeyboardInterrupt:
            driver.quit()
            print('用户主动中断爬虫，与现有浏览器连接断开')
            break

    # 重命名文件
    new_file_name = f"data/tb/淘宝_{urllib.parse.unquote(keyword)}_{time_string}_({record_num} of {total_num}).xlsx"
    try:
        os.rename(file_name, new_file_name)
        print(f"已将文件 {file_name} 重命名为 {new_file_name}")
    except Exception as e:
        print(e)
        print(f"重命名文件 {file_name} 失败")
    return [total_num, record_num]

def scrape_single_page(driver, keyword, start_page, page, file_name, headers):
    workbook = load_workbook(file_name)
    sheet = workbook.active
    last_row = sheet.max_row
    total_num = 0
    record_num = 0

    delay = random.uniform(min_delay, max_delay)
    time.sleep(delay)

    print("正在记录第"+str(page)+"页")
    try:
        if(page!=start_page):
            button = driver.find_element("xpath", "//button[contains(@class, 'next-btn') and contains(@class, 'next-medium') and contains(@class, 'next-btn-normal') and contains(@class, 'next-pagination-item') and contains(@class, 'next-next')]")
            button.click()
            time.sleep(2)
    except:
        print('点击下一页时出错')
    try:
        # 缓慢下拉页面
        scroll_height = driver.execute_script("return document.body.scrollHeight;")
        current_height = 0
        scroll_speed = 300  # 每次下拉的距离
        while current_height < scroll_height:
            driver.execute_script(f"window.scrollTo(0, {current_height});")
            current_height += scroll_speed
            time.sleep(0.3)  # 等待一段时间，模拟缓慢下拉的效果
            scroll_height = driver.execute_script("return document.body.scrollHeight;")
    except:
        print('下拉获取页面信息时发生错误')

    html = driver.execute_script(
        "return document.documentElement.outerHTML")
    # 创建 Beautiful Soup 对象
    soup = BeautifulSoup(html, "html.parser")
    # 使用 select 方法查找指定的元素
    elements = soup.select('div.Content--contentInner--QVTcU0M div a.Card--doubleCardWrapper--L2XFE73')
    print(len(elements))
    try:
        for (index, element) in enumerate(elements, start=1):
            goods_titles = element.select('div.Title--title--jCOPvpf')
            goods_nums = element.select('span.Price--realSales--FhTZc7U')
            goods_prices1 = element.select('span.Price--priceInt--ZlsSi_M')
            goods_prices2 = element.select('span.Price--priceFloat--h2RR0RK')
            total_num += 1
            # # 筛选
            # if (len(goods_titles) != 0):
            #     if filter_by_goods_name(goods_titles[0].text):
            #         continue
            if element.get('href').startswith('https:'):
                continue
            if convert_string_to_number(goods_nums[0].text.replace('人付款','').replace('人收货','')) < 200:
                continue

            record_num += 1

            # 下一行
            last_row+=1
            last_column = 0

            # 序号
            try:
                last_column+=1
                ordinal = last_row-1
                # sheet.column_dimensions[get_column_letter(last_column)].width = column_width
                # sheet.row_dimensions[last_row].height = row_height
                # ordinal_cell = sheet[f"{get_column_letter(last_column)}{last_row}"]
                # ordinal_cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
                sheet.cell(row=last_row, column=last_column, value=ordinal)
            except:
                print(f'记录“{headers[last_column-1]}”时出错')
                return

            # 电商平台
            try:
                last_column+=1
                platform_name = '淘宝'
                # sheet.column_dimensions[get_column_letter(last_column)].width = column_width
                # sheet.row_dimensions[last_row].height = row_height
                # current_time_cell = sheet[f"{get_column_letter(last_column)}{last_row}"]
                # current_time_cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
                sheet.cell(row=last_row, column=last_column, value=platform_name)
            except:
                print(f'记录“{headers[last_column-1]}”时出错')
                return

            # 关键词
            try:
                last_column+=1
                search_keyword = urllib.parse.unquote(keyword)
                # sheet.column_dimensions[get_column_letter(last_column)].width = column_width
                # sheet.row_dimensions[last_row].height = row_height
                # search_keyword_cell = sheet[f"{get_column_letter(last_column)}{last_row}"]
                # search_keyword_cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
                sheet.cell(row=last_row, column=last_column, value=search_keyword)
            except:
                print(f'记录“{headers[last_column-1]}”时出错')
                return

            # 店铺名称
            try:
                last_column+=1
                item = element.select('a.ShopInfo--shopName--rg6mGmy')
                if (len(item) != 0):
                    shop_name = ''
                    if has_unrecognized_characters(item[0].text):
                        shop_name = remove_unrecognized_characters(item[0].text)
                    else:
                        shop_name = item[0].text
                    # sheet.column_dimensions[get_column_letter(last_column)].width = column_width
                    # sheet.row_dimensions[last_row].height = row_height
                    # shop_name_cell = sheet[f"{get_column_letter(last_column)}{last_row}"]
                    # shop_name_cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
                    sheet.cell(row=last_row, column=last_column, value=shop_name)
                else:
                    sheet.cell(row=last_row, column=last_column, value='')
            except:
                print(f'记录“{headers[last_column-1]}”时出错')
                return

            # 店铺网址
            try:
                last_column+=1
                item = element.select('a.ShopInfo--shopName--rg6mGmy')
                if (len(item) != 0):
                    shop_link = item[0].get('href')
                    if not shop_link.startswith('https:'):
                        shop_link = 'https:' + shop_link
                    # sheet.column_dimensions[get_column_letter(last_column)].width = column_width
                    # sheet.row_dimensions[last_row].height = row_height
                    # shop_link_cell = sheet[f"{get_column_letter(last_column)}{last_row}"]
                    # shop_link_cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
                    # shop_link_cell.font = Font(underline="single", color="0563C1")
                    # shop_link_cell.hyperlink = shop_link
                    sheet.cell(row=last_row, column=last_column, value=shop_link)
                else:
                    sheet.cell(row=last_row, column=last_column, value='')
            except:
                print(f'记录“{headers[last_column-1]}”时出错')
                return
            
            # 店铺经营主体信息
            try:
                last_column+=1
                # sheet.column_dimensions[get_column_letter(last_column)].width = column_width
                # sheet.row_dimensions[last_row].height = row_height
                # manager_cell = sheet[f"{get_column_letter(last_column)}{last_row}"]
                # manager_cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
            except:
                print(f'记录“{headers[last_column-1]}”时出错')
                return

            # 商品图片
            try:
                last_column+=1
                item = element.select('img.MainPic--mainPic--rcLNaCv')
                if(len(item) != 0):
                    goods_img_url = item[0].get('src')
                    # sheet.column_dimensions[get_column_letter(last_column)].width = column_width
                    # sheet.row_dimensions[last_row].height = row_height
                    # goods_img_url_cell = sheet[f"{get_column_letter(last_column)}{last_row}"]
                    # goods_img_url_cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
                    # goods_img_url_cell.font = Font(underline="single", color="0563C1")
                    # goods_img_url_cell.hyperlink = goods_img_url
                    sheet.cell(row=last_row, column=last_column, value=goods_img_url)
            except:
                print(f'记录“{headers[last_column-1]}”时出错')
                return

            # 商品标题
            try:
                last_column+=1
                if (len(goods_titles) != 0):
                    goods_title = goods_titles[0].text.strip()
                    # sheet.column_dimensions[get_column_letter(last_column)].width = column_width
                    # sheet.row_dimensions[last_row].height = row_height
                    # shop_title_cell = sheet[f"{get_column_letter(last_column)}{last_row}"]
                    # shop_title_cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
                    sheet.cell(row=last_row, column=last_column, value=goods_title)
                else:
                    sheet.cell(row=last_row, column=last_column, value='')
            except:
                print(f'记录“{headers[last_column-1]}”时出错')
                return

            # 商品品牌
            try:
                last_column+=1
                # sheet.column_dimensions[get_column_letter(last_column)].width = column_width
                # sheet.row_dimensions[last_row].height = row_height
                # goods_brand_cell = sheet[f"{get_column_letter(last_column)}{last_row}"]
                # goods_brand_cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
            except:
                print(f'记录“{headers[last_column-1]}”时出错')
                return

            # 商品链接
            try:
                last_column+=1
                goods_link = element.get('href')
                if not goods_link.startswith('https:'):
                    goods_link = 'https:' + goods_link
                # sheet.column_dimensions[get_column_letter(last_column)].width = column_width
                # sheet.row_dimensions[last_row].height = row_height
                # goods_link_cell = sheet[f"{get_column_letter(last_column)}{last_row}"]
                # goods_link_cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
                # goods_link_cell.font = Font(underline="single", color="0563C1")
                # goods_link_cell.hyperlink = goods_link
                sheet.cell(row=last_row, column=last_column, value=goods_link)
            except:
                print(f'记录“{headers[last_column-1]}”时出错')
                return

            # 单价
            try:
                last_column+=1
                if (len(goods_prices1) != 0 and len(goods_prices2) != 0):
                    goods_price = goods_prices1[0].text + goods_prices2[0].text
                    goods_price = goods_price.strip()
                    # sheet.column_dimensions[get_column_letter(last_column)].width = column_width
                    # sheet.row_dimensions[last_row].height = row_height
                    # goods_price_cell = sheet[f"{get_column_letter(last_column)}{last_row}"]
                    # goods_price_cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
                    sheet.cell(row=last_row, column=last_column, value=goods_price)
                else:
                    sheet.cell(row=last_row, column=last_column, value='')
            except:
                print(f'记录“{headers[last_column-1]}”时出错')
                return

            # 销售量
            try:
                last_column+=1
                if (len(goods_nums) != 0):
                    goods_num = goods_nums[0].text.replace('人付款','').replace('人收货','')
                    # sheet.column_dimensions[get_column_letter(last_column)].width = column_width
                    # sheet.row_dimensions[last_row].height = row_height
                    # goods_num_cell = sheet[f"{get_column_letter(last_column)}{last_row}"]
                    # goods_num_cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
                    sheet.cell(row=last_row, column=last_column, value=goods_num)
            except:
                print(f'记录“{headers[last_column-1]}”时出错')
                return
            
            # 商品评论数
            try:
                last_column+=1
                # sheet.column_dimensions[get_column_letter(last_column)].width = column_width
                # sheet.row_dimensions[last_row].height = row_height
                # goods_commit_cell = sheet[f"{get_column_letter(last_column)}{last_row}"]
                # goods_commit_cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
            except:
                print(f'记录“{headers[last_column-1]}”时出错')
                return

            # 销售额
            try:
                last_column+=1
                if(len(goods_prices1) != 0 and len(goods_prices2) != 0 and len(goods_nums) != 0):
                    goods_price = is_float(goods_prices1[0].text + goods_prices2[0].text) and float(goods_prices1[0].text + goods_prices2[0].text) or 0
                    goods_num = convert_string_to_number(goods_nums[0].text.replace('人付款','').replace('人收货',''))
                    goods_sales = goods_price * goods_num
                    # sheet.column_dimensions[get_column_letter(last_column)].width = column_width
                    # sheet.row_dimensions[last_row].height = row_height
                    # goods_sales_cell = sheet[f"{get_column_letter(last_column)}{last_row}"]
                    # goods_sales_cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
                    sheet.cell(row=last_row, column=last_column, value=goods_sales)
                else:
                    sheet.cell(row=last_row, column=last_column, value='0')
            except:
                print(f'记录“{headers[last_column-1]}”时出错')
                return
    except Exception as e:
        print('表格记录数据时出错')
        driver.quit()
        print('与现有浏览器连接断开')

    try:
        workbook.save(file_name)
        print(f"已保存第 {page} 页数据到 {file_name}")
    except Exception as e:
        print(e)
        driver.quit()
        print('与现有浏览器连接断开')
    return [total_num, record_num]

def filter_by_shop_name(shopName):
    keywords = ['华为京东自营官方旗舰店']
    if shopName in keywords:
        return True
    else:
        return False

def filter_by_goods_name(goodsName):
    required_keywords = 3  # 至少需要满足的关键字数
    lower_case_good_name = goodsName.lower()
    matched_keywords = 0
    keywords = ['xiaomi', 'huawei', 'oppo', 'vivo', 'redmi', 'realme', '真我', '红米', '小米', '华为', '荣耀', '魅族', '一加', '苹果']
    for keyword in keywords:
        lower_case_keyword = keyword.lower()
        if lower_case_keyword in lower_case_good_name:
            matched_keywords += 1
            if matched_keywords >= required_keywords:
                return True
    return False

def convert_string_to_number(string):
    if not string:
        return 0
    if string.endswith('万+'):
        number = int(string[:-2]) * 10000
    elif string.endswith('+'):
        number = int(string[:-1])
    else:
        number = int(string)
    return number

def is_float(string):
    try:
        float(string)
        return True
    except ValueError:
        return False

def has_unrecognized_characters(string):
    return not all(char.isprintable() for char in string)

def remove_unrecognized_characters(string):
    return ''.join(char for char in string if char.isprintable())

if __name__ == "__main__":
    # options = webdriver.FirefoxOptions()
    # driver = webdriver.Remote(
    #     command_executor="http://127.0.0.1:4444", options=options)
    # 创建带有Selenium Wire的Firefox WebDriver对象
    # options = webdriver.FirefoxOptions()
    # options.set_preference('network.proxy.type', 1)
    # options.set_preference('network.proxy.http', 'localhost')
    # options.set_preference('network.proxy.http_port', 8888)
    # driver = webdriver.Firefox(options=options)

    # options = webdriver.FirefoxOptions()
    # options.add_argument("--disable-blink-features=AutomationControlled")
    # options.add_argument("--headless")
    # driver = webdriver.Remote(
    #     command_executor="http://127.0.0.1:4444", options=options)
    # driver.maximize_window()
    driver = webdriver.Firefox()
    login(driver)
    print("登录成功")

    # url = 'https://s.taobao.com/search'
    url = input('请输入网址：')
    driver.get(url)

    keywords = ['华为原厂+华为4k','海思芯+华为4k']
    for keyword in  keywords:
        keyword = urllib.parse.quote(keyword)
        start_page = 1
        end_page = 100

        [total_num, record_num] = scrape_multiple_pages(driver, keyword, start_page, end_page)
        print(f"共找到 {total_num} 条数据，经过筛选，已记录 {record_num} 条数据")
        
    driver.quit()
    print('与现有浏览器连接断开')