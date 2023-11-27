from selenium import webdriver
from openpyxl import load_workbook
from bs4 import BeautifulSoup
import time
import re
import io
from PIL import Image
from datetime import datetime
from wsgiref.handlers import format_date_time
from time import mktime
import hashlib
import base64
import hmac
from urllib.parse import urlencode
import json
import requests


APPId = "b9b55ac3"  # 控制台获取
APISecret = "NzA4N2QxMDNlZjQzMzk5YjUzYzM5NzM3"  # 控制台获取
APIKey = "602f362334c3f47d5c626652e91f25dc"  # 控制台获取

file_name = 'data/jd/merge/需求2_京东.xlsx'
row_height = 40
column_width = 14

# 打开需读取的excel表
workbook = load_workbook(file_name)
sheet = workbook.active

# 打开火狐浏览器模拟器
options = webdriver.FirefoxOptions()
driver = webdriver.Remote(command_executor="http://127.0.0.1:4444", options=options)

# options = webdriver.FirefoxOptions()
# driver = webdriver.Firefox(options=options)

class AssembleHeaderException(Exception):
    def __init__(self, msg):
        self.message = msg

class Url:
    def __init__(self, host, path, schema):
        self.host = host
        self.path = path
        self.schema = schema
        pass

# calculate sha256 and encode to base64
def sha256base64(data):
    sha256 = hashlib.sha256()
    sha256.update(data)
    digest = base64.b64encode(sha256.digest()).decode(encoding='utf-8')
    return digest

def parse_url(requset_url):
    stidx = requset_url.index("://")
    host = requset_url[stidx + 3:]
    schema = requset_url[:stidx + 3]
    edidx = host.index("/")
    if edidx <= 0:
        raise AssembleHeaderException("invalid request url:" + requset_url)
    path = host[edidx:]
    host = host[:edidx]
    u = Url(host, path, schema)
    return u

# build websocket auth request url
def assemble_ws_auth_url(requset_url, method="POST", api_key="", api_secret=""):
    u = parse_url(requset_url)
    host = u.host
    path = u.path
    now = datetime.now()
    date = format_date_time(mktime(now.timetuple()))
    # date = "Thu, 12 Dec 2019 01:57:27 GMT"
    signature_origin = "host: {}\ndate: {}\n{} {} HTTP/1.1".format(host, date, method, path)
    signature_sha = hmac.new(api_secret.encode('utf-8'), signature_origin.encode('utf-8'),
                             digestmod=hashlib.sha256).digest()
    signature_sha = base64.b64encode(signature_sha).decode(encoding='utf-8')
    authorization_origin = "api_key=\"%s\", algorithm=\"%s\", headers=\"%s\", signature=\"%s\"" % (
        api_key, "hmac-sha256", "host date request-line", signature_sha)
    authorization = base64.b64encode(authorization_origin.encode('utf-8')).decode(encoding='utf-8')
    values = {
        "host": host,
        "date": date,
        "authorization": authorization
    }

    return requset_url + "?" + urlencode(values)

# 分类
try:
    dict = {}
    start_row = 2
    end_row = sheet.max_row

    total = end_row - start_row + 1
    current = 0
    start_time = time.time()
    time.sleep(1)
    print(f'\n正在分类')
    for row in range(start_row, end_row + 1):
        shop_id = re.findall(r'\d+', sheet.cell(row=row, column=5).value)[0]
        value = sheet.cell(row=row, column=6).value
        if value == None:
            if shop_id in dict:
                dict[shop_id].append(row)
            else:
                dict[shop_id] = [row]
except Exception as e:
    driver.quit()
    print(e)
    print('分类时出错')

# 爬取店铺经营信息
try:
    start_row = 2
    end_row = sheet.max_row

    total = len(dict)
    current = 0
    start_time = time.time()
    time.sleep(1)
    print(f'\n正在爬取店铺经营信息')
    for key, val in dict.items():
        current+=1
        res = (total - current) / (current / ((time.time() - start_time) / 60))
        print(f"\r当前进度：{current}/{total}，预计仍需：{res:.2f} min", end="")
        url = 'https://mall.jd.com/showLicence-'+key+'.html'
        driver.get(url)
        finish = False
        while not finish:
            tempHTML = driver.execute_script("return document.documentElement.outerHTML")
            tempSoup = BeautifulSoup(tempHTML, "html.parser")
            verifyCodeImg = tempSoup.find_all('img',id='verifyCodeImg')
            if len(verifyCodeImg) != 0:
                try:
                    # image_url = 'https:'+verifyCodeImg[0].get('src')  # 替换为你要识别的图片的URL
                    # print(image_url)
                    # temp_image_path = 'temp/temp.jpg'
                    # response = requests.get(image_url)
                    # with open(temp_image_path, 'wb') as f:
                    #     f.write(response.content)

                    verifyCodeImg = driver.find_element("xpath","//img[@id='verifyCodeImg']")
                    # 获取元素的位置和大小
                    element_location = verifyCodeImg.location
                    element_size = verifyCodeImg.size

                    # 获取整个页面的截图
                    screenshot = driver.get_screenshot_as_png()
                    # 将截图转换为 Image 对象
                    image = Image.open(io.BytesIO(screenshot))

                    # 计算指定元素的区域
                    left = element_location['x'] + 147
                    top = element_location['y'] + 85
                    right = element_location['x'] + element_size['width'] + 173
                    bottom = element_location['y'] + element_size['height'] + 95
                    element_region = (left, top, right, bottom)

                    # 裁剪指定元素的区域
                    element_screenshot = image.crop(element_region)

                    # 保存截图
                    element_screenshot.save('temp/element_screenshot.png')
                    temp_image_path = 'temp/element_screenshot.png'

                    with open(temp_image_path, "rb") as f:
                        imageBytes = f.read()

                    url = 'https://api.xf-yun.com/v1/private/sf8e6aca1'

                    body = {
                        "header": {
                            "app_id": APPId,
                            "status": 3
                        },
                        "parameter": {
                            "sf8e6aca1": {
                                "category": "ch_en_public_cloud",
                                "result": {
                                    "encoding": "utf8",
                                    "compress": "raw",
                                    "format": "json"
                                }
                            }
                        },
                        "payload": {
                            "sf8e6aca1_data_1": {
                                "encoding": "jpg",
                                "image": str(base64.b64encode(imageBytes), 'UTF-8'),
                                "status": 3
                            }
                        }
                    }

                    request_url = assemble_ws_auth_url(url, "POST", APIKey, APISecret)
                    headers = {'content-type': "application/json", 'host': 'api.xf-yun.com', 'app_id': APPId}
                    response = requests.post(request_url, data=json.dumps(body), headers=headers)
                    tempResult = json.loads(response.content.decode())
                    finalResult = base64.b64decode(tempResult['payload']['result']['text']).decode()
                    finalResult = json.loads(finalResult.replace(" ", "").replace("\n", "").replace("\t", "").strip())    
                    try:
                        code = finalResult['pages'][0]['lines'][0]['words'][0]['content'].replace('-', '').replace('￥', 'Y').replace('(', 'C').replace('（', 'C').replace('+', 't')
                    except:
                        print('遇到验证码')
                        driver.get(url)
                        time.sleep(5)
                        continue
                    verifyCodeInput = driver.find_element("xpath","//input[@id='verifyCode']")
                    driver.execute_script("arguments[0].setAttribute('autocomplete', 'off')", verifyCodeInput)
                    verifyCodeInput.send_keys(code)
                    time.sleep(2)
                    sutmit = driver.find_element("xpath","//button[contains(@class, 'btn') and @type='submit']")
                    sutmit.click()
                    time.sleep(1)

                    tempHTML = driver.execute_script("return document.documentElement.outerHTML")
                    tempSoup = BeautifulSoup(tempHTML, "html.parser")
                    qualificationItem = tempSoup.select('li.qualification-item')
                    messageError = tempSoup.select('li.message_error')
                    if len(qualificationItem) != 0:
                        finish = True
                        tempHTML = driver.execute_script("return document.documentElement.outerHTML")
                        tempSoup = BeautifulSoup(tempHTML, "html.parser")
                        elements = tempSoup.select('li.noBorder span')
                        for row in val:
                            sheet.cell(row=row, column=6, value=elements[0].text)
                    elif len(messageError) == 0:
                        finish = True
                        for row in val:
                            value = sheet.cell(row=row, column=4).value
                            sheet.cell(row=row, column=6, value=value + '：该店铺无经营企业名称')
                except Exception as e:
                    print(e)
            else:
                finish = True
                for row in val:
                    value = sheet.cell(row=row, column=4).value
                    sheet.cell(row=row, column=6, value=value + '：该店铺无证照')
except Exception as e:
    print(e)
    print('爬取店铺经营信息时出错')
finally:
    workbook.save(file_name)
    driver.quit()
    print('与现有浏览器连接断开')
    end_time = time.time()
    duration = end_time - start_time
    print(f"爬虫耗时：{duration:.2f} 秒")
    print(f"目标数量：{total} 条")
    print(f"已获取数量：{current} 条")
    unit = current / (duration / 60)
    print(f"每分钟爬取数量：{unit:.2f} 条")