# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from openpyxl import Workbook
import time
import os
from openpyxl.utils.cell import get_column_letter
import re
import json

file_name = "data/jd/merge/1628.xlsx"
num = 3
new_file_name = file_name.replace('.xlsx','_') + str(num) + '.xlsx'
json_path = "src/jd/data_files/filter.json"

# 打开需读取的excel表
workbook = load_workbook(file_name)
sheet = workbook.active

# 新建excel表
new_workbook = Workbook()
new_sheet = new_workbook.active

# 处理表头
print(f'\n正在处理表头')
first_row = sheet[1]
for cell in first_row:
    new_sheet[cell.coordinate].value = cell.value

# 通过店铺白名单筛选
try:
    def convert_string_to_number(string):
        if not string:
            return 0
        if isinstance(string, int):
            return string
        if string.endswith('万+'):
            number = int(string[:-2]) * 10000
        elif string.endswith('+'):
            number = int(string[:-1])
        else:
            number = int(string)
        return number
    def check_keywords1(text):
        keywords = ['耳机']
        pattern = '|'.join(keywords)
        match = re.search(pattern, text, flags=re.IGNORECASE)
        return match is not None
    def check_keywords2(text):
        keywords = ['HUAWEI','华为']
        pattern = '|'.join(keywords)
        match = re.search(pattern, text, flags=re.IGNORECASE)
        return match is not None
    list = ['一号会员店','JDG官方旗舰店','丰弘数码专营店','华为（HUAWEI）企业业务京东自营官方旗舰店','华为（HUAWEI）数码京喜直营专区','华为次第华开专卖店','华为京东自营官方旗舰店','华为礼象专卖店','华为秒通专卖店','京东电竞手机官方旗舰店','京东通信京东自营旗舰店','京东手机运营商自营旗舰店','京东手机自营旗舰店','商用电脑办公设备京东自营专区','商用电脑数码终端设备京东自营专区','上海电信京东自营旗舰店','上海移动京东自营官方旗舰店','天翼电信京东自营旗舰店','浙江电信京东自营旗舰店','中国电信京东自营旗舰店','倍思（Baseus）车品京东自营旗舰店','倍思旗舰店','倍思（Baseus）京东自营旗舰店','罗马仕京东自营旗舰店','万魔官方旗舰店','迪士尼（DISNEY）手机配件京东自营专区','公牛电器官方旗舰店','酷狗京东自营旗舰店','公牛电器官方旗舰店','闪魔京东自营旗舰店','闪魔旗舰店','万魔官方旗舰店']
    # 读取现有店铺信息
    with open(json_path, encoding='utf-8') as file:
        other_list = json.load(file)
    record_list = []
    start_row = 2
    end_row = sheet.max_row

    total = end_row - start_row + 1
    current = 0
    start_time = time.time()
    time.sleep(1)
    print(f'\n正在通过店铺白名单筛选')
    for row in range(start_row, end_row + 1):
        shop_name = sheet.cell(row=row, column=4).value
        goods_title = sheet.cell(row=row, column=8).value
        goods_nums = sheet.cell(row=row, column=13).value
        if convert_string_to_number(goods_nums) >= 200 and shop_name not in list and shop_name not in other_list and check_keywords1(goods_title) and check_keywords2(goods_title):
            record_list.append(row)
        # if sheet.cell(row=row, column=15).value != 'delete':
        #     record_list.append(row)
except Exception as e:
    print(e)
    print('通过店铺白名单筛选时出错')

# 记录数据到新表
try:
    start_row = 2
    end_row = sheet.max_row

    total = len(record_list)
    current = 0
    start_time = time.time()
    time.sleep(1)
    print(f'\n正在记录数据到新表')
    for row in record_list:
        current+=1
        res = (total - current) / (current / ((time.time() - start_time) / 60))
        print(f"\r当前进度：{current}/{total}，预计仍需：{res:.2f} min", end="")
        for cell in sheet[row]:
            new_sheet[f"{get_column_letter(cell.column)}{current+1}"].value = cell.value
except Exception as e:
    print(e)
    print('记录数据到新表时出错')

# 处理序号
try:
    start_row = 2
    end_row = new_sheet.max_row

    total = total
    current = 0
    start_time = time.time()
    time.sleep(1)
    print(f'\n正在处理序号')
    for row in range(start_row, end_row + 1):
        current+=1
        res = (total - current) / (current / ((time.time() - start_time) / 60))
        print(f"\r当前进度：{current}/{total}，预计仍需：{res:.2f} min", end="")
        new_sheet.cell(row=row, column=1, value=row-1)
except Exception as e:
    print(e)
    print('处理序号时出错')

new_workbook.save(new_file_name)

# # 修改文件名
# try:
#     temp_file_name = "/".join(file_name.split("/")[:-1]) + '/temp.xlsx'
#     os.rename(file_name, temp_file_name)
#     os.rename(new_file_name, file_name)
#     os.rename(temp_file_name, new_file_name)
# except Exception as e:
#     print(e)
#     print('修改文件名时出错')