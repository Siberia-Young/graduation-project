from selenium import webdriver
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import time
from selenium.webdriver.common.proxy import Proxy, ProxyType
import shutil
import cv2

file_name = "data/jd/merge/outcome/文件5.2.xlsx"

# try:
#     copy_file_name = file_name.replace('.xlsx','(副本).xlsx')
#     shutil.copy(file_name, copy_file_name)
#     temp_workbook = load_workbook(copy_file_name)
#     temp_sheet = temp_workbook.active
#     for row in range(2, temp_sheet.max_row+1):
#         temp_sheet.cell(row=row, column=16, value='')
#     temp_workbook.save(copy_file_name)
# except:
#     print(f'\n出错')

workbook = load_workbook(file_name)
sheet = workbook.active
start_time = time.time()

start_row = 2
end_row = sheet.max_row

total = end_row - start_row + 1
current = 0

# 创建代理对象
# proxy = Proxy()
# proxy.proxy_type = ProxyType.MANUAL
# proxy.http_proxy = '183.7.128.113:45151'
# 打开火狐浏览器模拟器
options = webdriver.FirefoxOptions()
# options.add_argument('--proxy-server={}'.format(proxy.http_proxy))
driver = webdriver.Remote(command_executor="http://127.0.0.1:4444", options=options)

# options = webdriver.FirefoxOptions()
# driver = webdriver.Firefox(options=options)

try:
    def image_match(base_path, target_path):
        base_img = cv2.imread(base_path)
        target_img = cv2.imread(target_path)
        result = cv2.matchTemplate(target_img, base_img, cv2.TM_CCOEFF_NORMED)
        min_val, max_val, min_loc, max_loc = cv2.minMaxLoc(result)
        print(max_val)
        if (max_val >= 0.99):
            return True
        return False
    
    count = 0
    end = False
    while count != 0 or end == False:
        count = 0
        if end == True:
            current = 0
            print(f'\n新一轮处理')
        for row in range(start_row, end_row + 1):
            current+=1
            res = (total - current) / (current / ((time.time() - start_time) / 60))
            print(f"\r当前进度：{current}/{total}，预计仍需：{res:.2f} min", end="")

            # 如果品牌字段不为空，且该行被标记应该被删除，这跳过不处理
            temp = sheet.cell(row=row, column=9).value
            delete = sheet.cell(row=row, column=15).value
            if (temp != None) or (delete != None and delete == 'delete'):
                continue

            goods_link = sheet.cell(row=row, column=10).value
            driver.get(goods_link)

            try:
                # 缓慢下拉页面
                scroll_height = driver.execute_script("return document.body.scrollHeight;")
                scroll_count = 0
                current_height = 0
                scroll_speed = 500  # 每次下拉的距离
                while scroll_count < 3:
                    driver.execute_script(f"window.scrollTo(0, {current_height});")
                    current_height += scroll_speed
                    scroll_count += 1
                    time.sleep(0.1)  # 等待一段时间，模拟缓慢下拉的效果
                time.sleep(1)
            except:
                print('下拉获取页面信息时发生错误')

            tempHTML = driver.execute_script("return document.documentElement.outerHTML")
            tempSoup = BeautifulSoup(tempHTML, "html.parser")

            # 遇到登录页面或者验证页面跳过，等下一轮再处理
            login = tempSoup.select('div.login-btn')
            verify = tempSoup.select('div.verifyBtn')
            if len(login) != 0 or len(verify) != 0:
                continue

            elements = tempSoup.select('div.hxm_hide_page')
            if len(elements) == 0:
                elements = tempSoup.select('div.itemover-tip')
            if len(elements) == 0:
                elements = tempSoup.select('div.logo_extend')
            if len(elements) == 0:
                try:
                    goods_brand_element = tempSoup.find_all('ul',id='parameter-brand')
                    if len(goods_brand_element) != 0:
                        goods_brand = goods_brand_element[0].select('li')[0].text
                        if goods_brand.startswith('品牌：'):
                            goods_brand = goods_brand.replace('品牌：', '').replace('\n', '').replace('\r', '').replace(' ', '')
                        else:
                            goods_brand = "暂无"
                    else:
                        goods_brand = "暂无"
                    
                    sheet.cell(row=row, column=9, value=goods_brand)
                    
                    choose = tempSoup.select('div.li.p-choose:not(.hide)')
                    choose_text_list = []
                    for item in choose:
                        choose_list = item.select('div.dd div a')
                        for item1 in choose_list:
                            choose_text_list.append(item1.text.strip())
                    sheet.cell(row=row, column=17, value='\n'.join(choose_text_list))

                    parameter = tempSoup.select('div.p-parameter ul.p-parameter-list')
                    parameter_text_list = []
                    for item in parameter:
                        parameter_list = item.select('li')
                        for item1 in parameter_list:
                            parameter_text_list.append(item1.text.strip())
                    sheet.cell(row=row, column=18, value='\n'.join(parameter_text_list))

                    more_imgs = tempSoup.select('div.spec-items ul.lh li img')
                    if len(more_imgs) != 0:
                        imgs_list = []
                        for img in more_imgs:
                            img_src = 'https:' + img.get('src').replace('.avif','').replace('/n5/','/n1/')
                            imgs_list.append(img_src)
                        sheet.cell(row=row, column=19, value='\n'.join(imgs_list))
                    
                    
                    
                    # if sheet.cell(row=row, column=4).value is None:
                    #     shop_element = tempSoup.select('div.popbox-inner h3 a')
                        
                    #     shop_name = len(shop_element) == 0 and '暂无' or shop_element[0].text
                    #     shop_name_cell = sheet[f"{get_column_letter(4)}{row}"]
                    #     shop_name_cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
                    #     sheet.cell(row=row, column=4, value=shop_name)

                    #     shop_link = len(shop_element) == 0 and '暂无' or ('https:'+shop_element[0].get('href'))
                    #     shop_link_cell = sheet[f"{get_column_letter(5)}{row}"]
                    #     shop_link_cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
                    #     shop_link_cell.font = Font(underline="single", color="0563C1")
                    #     shop_link_cell.hyperlink = shop_link
                    #     sheet.cell(row=row, column=5, value=shop_link)
                except:
                    workbook.save(file_name)
                    driver.quit()
                    print('与现有浏览器连接断开')
            else:
                sheet.cell(row=row, column=15, value='delete')
            count += 1
        workbook.save(file_name)
        end = True
except Exception as e:
    print(e)
    print('主动中断')
finally:
    # 保存文件
    workbook.save(file_name)
    driver.quit()
    print('与现有浏览器连接断开')
    end_time = time.time()
    duration = end_time - start_time
    print(f"爬虫耗时：{duration:.2f} 秒")
    print(f"目标数量：{total} 条")
    print(f"已获取数量：{current} 条")
    unit = current / (duration / 60)
    print(f"每分钟爬取数量：{unit:.2f} 条")
