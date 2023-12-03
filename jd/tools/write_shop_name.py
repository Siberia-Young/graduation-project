from openpyxl import load_workbook
from openpyxl import Workbook
import time
import os
from openpyxl.utils.cell import get_column_letter
import json

file_name = "data/jd/2023-11-29/jd_2023-11-29/文件4_(2096).xlsx"
num = 5

json_path = "src/jd/data_files/filter.json"
# 打开需读取的excel表
workbook = load_workbook(file_name)
sheet = workbook.active

# 读取现有店铺信息
with open(json_path, encoding='utf-8') as file:
    list = json.load(file)

# 记录数据到json
try:
    start_row = 2
    end_row = sheet.max_row

    total = end_row - start_row + 1
    current = 0
    start_time = time.time()
    time.sleep(1)
    print(f'\n正在记录数据到json')
    for row in range(start_row, end_row + 1):
        current+=1
        res = (total - current) / (current / ((time.time() - start_time) / 60))
        print(f"\r当前进度：{current}/{total}，预计仍需：{res:.2f} min", end="")
        value = sheet.cell(row=row, column=4).value
        if value not in list:
            list.append(value)
except Exception as e:
    print(e)
finally:
   with open(json_path, 'w', encoding='utf-8') as file:
        json.dump(list, file, indent=4, ensure_ascii=False)