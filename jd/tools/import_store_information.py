from aip import AipOcr
import ssl
from selenium import webdriver
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from bs4 import BeautifulSoup
import time
from os import path
from aip import AipOcr
import re
import requests
import os
import io
from PIL import Image
import json


file_name = "data/jd/merge/merge_2_3.xlsx"
num = 2
base_file_name = 'src/jd/data_files/shop_body_info/shop_body_info.json'
row_height = 40
column_width = 14

def import_store_information(file_name,base_file_name = 'src/jd/data_files/shop_body_info/shop_body_info.json'):
    # 打开需读取的excel表
    workbook = load_workbook(file_name)
    sheet = workbook.active

    # 分类
    try:
        record_dict = {}
        start_row = 2
        end_row = sheet.max_row

        total = end_row - start_row + 1
        current = 0
        start_time = time.time()
        time.sleep(1)
        print(f'\n正在分类')
        for row in range(start_row, end_row + 1):
            shop_link = sheet.cell(row=row, column=5).value
            shop_body = sheet.cell(row=row, column=6).value
            if shop_body == None:
                if shop_link in record_dict:
                    record_dict[shop_link].append(row)
                else:
                    record_dict[shop_link] = [row]
    except Exception as e:
        print(e)
        print('分类时出错')

    # 读取现有店铺信息
    try:
        print(f'\n正在读取现有店铺信息')
        with open(base_file_name, encoding='utf-8') as file:
            dict = json.load(file)
    except Exception as e:
        print(e)
        print('读取现有店铺信息时出错')

    # 记录需要导入的店铺列表
    try:
        print(f'\n正在记录需要导入的店铺列表')
        import_list = []
        for key, val in record_dict.items():
            if key in dict:
                import_list.append([key, val])
    except Exception as e:
        print(e)
        print('记录需要导入的店铺列表时出错')

    # 导入经营信息
    try:
        start_row = 2
        end_row = sheet.max_row

        total = len(import_list)
        current = 0
        start_time = time.time()
        time.sleep(1)
        print(f'\n正在导入经营信息')
        for item in import_list:
            current+=1
            res = (total - current) / (current / ((time.time() - start_time) / 60))
            print(f"\r当前进度：{current}/{total}，预计仍需：{res:.2f} min", end="")
            for row in item[1]:
                sheet.cell(row=row, column=6, value=dict[item[0]][1])
    except Exception as e:
        print(e)
        print('导入经营信息时出错')
    finally:
        workbook.save(file_name)
        end_time = time.time()
        duration = end_time - start_time
        print(f"导入耗时：{duration:.2f} 秒")
        print(f"目标数量：{total} 条")
        print(f"已导入数量：{current} 条")
        unit = current / (duration / 60)
        print(f"每分钟导入数量：{unit:.2f} 条")

# import_store_information(file_name)