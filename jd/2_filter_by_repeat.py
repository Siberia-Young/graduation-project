from openpyxl import load_workbook
from openpyxl import Workbook
import time
import os

file_name = "data/jd/merge/merge.xlsx"
num = 2
new_file_name = file_name.replace('.xlsx','_') + str(num) + '.xlsx'

# 打开需读取的excel表
workbook = load_workbook(file_name)
sheet = workbook.active

# 新建excel表
new_workbook = Workbook()
new_sheet = new_workbook.active

# 处理表头
print(f'\n正在处理表头')
first_row = sheet[1]
for cell in first_row:
    new_sheet[cell.coordinate].value = cell.value

# 通过商品链接进行筛选去重并记录到新excel表
try:
    list = []
    start_row = 2
    end_row = sheet.max_row

    total = end_row - start_row + 1
    current = 0
    start_time = time.time()
    time.sleep(1)
    print(f'\n正在比对去重并记录到新的excel表')
    for row in range(start_row, end_row + 1):
        current+=1
        res = (total - current) / (current / ((time.time() - start_time) / 60))
        print(f"\r当前进度：{current}/{total}，预计仍需：{res:.2f} min", end="")
        value = sheet.cell(row=row, column=10).value
        if value not in list:
            list.append(value)
            for cell in sheet[row]:
                new_sheet[cell.coordinate].value = cell.value
except Exception as e:
    print(e)
    print('通过商品链接进行筛选去重并记录到新excel表时出错')

# 删除新excel表空白行
try:
    list = []
    print(f'\n正在删除空白行')
    for row in range(start_row, end_row + 1):
        value = new_sheet.cell(row=row, column=1).value
        if(value==None):
            list.append(row)

    total = len(list)
    current = 0
    start_time = time.time()
    time.sleep(1)
    for row in reversed(list):
        current+=1
        res = (total - current) / (current / ((time.time() - start_time) / 60))
        print(f"\r当前进度：{current}/{total}，预计仍需：{res:.2f} min", end="")
        new_sheet.delete_rows(row)
except Exception as e:
    print(e)
    print('删除新excel表空白行时出错')

# 处理序号
try:
    start_row = 2
    end_row = new_sheet.max_row

    total = end_row - start_row + 1
    current = 0
    start_time = time.time()
    time.sleep(1)
    print(f'\n正在处理序号')
    for row in range(start_row, end_row + 1):
        current+=1
        res = (total - current) / (current / ((time.time() - start_time) / 60))
        print(f"\r当前进度：{current}/{total}，预计仍需：{res:.2f} min", end="")
        new_sheet.cell(row=row, column=1, value=row-1)
except Exception as e:
    print(e)
    print('处理序号时出错')

new_workbook.save(new_file_name)

# 修改文件名
try:
    temp_file_name = "/".join(file_name.split("/")[:-1]) + '/temp.xlsx'
    os.rename(file_name, temp_file_name)
    os.rename(new_file_name, file_name)
    os.rename(temp_file_name, new_file_name)
except Exception as e:
    print(e)
    print('修改文件名时出错')