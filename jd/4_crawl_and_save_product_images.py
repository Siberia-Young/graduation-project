from openpyxl import load_workbook
import time
import requests
import os

file_name = "data/pdd/拼多多_华为手表_2023-11-06_18-32-40.xlsx"
num = 4
folder_path = "/".join(file_name.split("/")[:-1]) + '/images'

# 打开需读取的excel表
workbook = load_workbook(file_name)
sheet = workbook.active

# 处理文件夹不存在的情况
if not os.path.exists(folder_path):
    os.makedirs(folder_path)

# 下载图片
try:
    fail_list = []
    start_row = 2
    end_row = sheet.max_row

    total = end_row - start_row + 1
    current = 0
    start_time = time.time()
    time.sleep(1)
    print(f'\n正在下载图片')
    for row in range(start_row, end_row + 1):
        current+=1
        res = (total - current) / (current / ((time.time() - start_time) / 60))
        print(f"\r当前进度：{current}/{total}，预计仍需：{res:.2f} min", end="")
        value = sheet.cell(row=row, column=7).value
        response = requests.get(value)
        if response.status_code == 200:
            image_path = os.path.join(folder_path, f'{row}.{value.split(".")[-1]}')
            with open(image_path, 'wb') as file:
                file.write(response.content)
        else:
            fail_list.append(row)
except:
    print("关闭VPN")
finally:
    end_time = time.time()
    duration = end_time - start_time
    print(f"下载耗时：{(duration/60):.2f} min")
    print(f"目标数量：{total} 条")
    print(f"已下载数量：{current} 条")
    unit = current / (duration / 60)
    print(f"每分钟下载数量：{unit:.2f} 条")

print(f'下载失败：{fail_list}')