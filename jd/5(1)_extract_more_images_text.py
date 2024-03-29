import easyocr
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import time
import os
import shutil

file_name = "data/jd/merge/merge_2_3_8_9_6_9.xlsx"
num = 5
folder_path = "/".join(file_name.split("/")[:-1]) + '/more_images'

try:
    copy_file_name = file_name.replace('.xlsx','(副本).xlsx')
    shutil.copy(file_name, copy_file_name)
    temp_workbook = load_workbook(copy_file_name)
    temp_sheet = temp_workbook.active
    for row in range(2, temp_sheet.max_row+1):
        temp_sheet.cell(row=row, column=16, value='')
    temp_workbook.save(copy_file_name)
except:
    print(f'\n出错')

# 打开需读取的excel表
workbook = load_workbook(file_name)
sheet = workbook.active

# 只需要运行一次就可以将模型加载到内存中
reader = easyocr.Reader(['ch_sim','en'])

# 识别图片提取文字
try:
    start_row = 2
    end_row = sheet.max_row

    total = end_row - start_row + 1
    current = 0
    start_time = time.time()
    time.sleep(1)
    print(f'\n正在识别图片提取文字')
    for row in range(start_row, end_row + 1):
        current+=1
        res = (total - current) / (current / ((time.time() - start_time) / 60))
        print(f"\r当前进度：{current}/{total}，预计仍需：{res:.2f} min", end="")
        imgs = sheet.cell(row=row, column=19).value
        if sheet.cell(row=row, column=20).value != None:
            continue
        if imgs == None:
            continue
        result_list = []
        for index,img in enumerate(imgs.split('\n')):
            image_path = os.path.join(folder_path, f'{row}.{index}.{img.split(".")[-1]}')
            if os.path.exists(image_path):
                try:
                    result = reader.readtext(image_path, detail = 0, paragraph=True)
                    result_list.append(' '.join(result))
                except Exception as e:
                    print(e)
                    print(f'识别出错：{row}.{index}')
            else:
                print(f'找不到图片路径：{row}.{index}')
        sheet.cell(row=row, column=20, value=' '.join(result_list))
except Exception as e:
    print(e)
    print('出错')
finally:
    workbook.save(file_name)
    end_time = time.time()
    duration = end_time - start_time
    print(f"识别耗时：{(duration/60):.2f} min")
    print(f"目标数量：{total} 条")
    print(f"已识别数量：{current} 条")
    unit = current / (duration / 60)
    print(f"每分钟识别数量：{unit:.2f} 条")