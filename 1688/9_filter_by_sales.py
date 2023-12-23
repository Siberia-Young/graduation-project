from openpyxl import load_workbook
from openpyxl import Workbook
import time
import os
from openpyxl.utils.cell import get_column_letter

file_name = "data/1688/merge/merge_2_3_8.xlsx"
num = 9
new_file_name = file_name.replace('.xlsx','_') + str(num) + '.xlsx'

# 打开需读取的excel表
workbook = load_workbook(file_name)
sheet = workbook.active

# 新建excel表
new_workbook = Workbook()
new_sheet = new_workbook.active

# 处理表头
print(f'\n正在处理表头')
first_row = sheet[1]
for cell in first_row:
    new_sheet[cell.coordinate].value = cell.value

# 分类
try:
    dict = {}
    start_row = 2
    end_row = sheet.max_row

    total = end_row - start_row + 1
    current = 0
    start_time = time.time()
    time.sleep(1)
    print(f'\n正在分类')
    for row in range(start_row, end_row + 1):
        current+=1
        res = (total - current) / (current / ((time.time() - start_time) / 60))
        print(f"\r当前进度：{current}/{total}，预计仍需：{res:.2f} min", end="")
        value = sheet.cell(row=row, column=5).value
        if value in dict:
            dict[value].append(row)
        else:
            dict[value] = [row]
except Exception as e:
    print(e)
    print('分类时出错')

# 通过店铺内商品总销售额筛选并记录到新表
try:
    start_row = 2
    end_row = sheet.max_row

    total = 0
    keys_to_delete = []
    for key, val in dict.items():
        sum = 0
        for row in val:
            value = sheet.cell(row=row, column=14).value
            sum += value
        if sum < 500000:
            keys_to_delete.append(key)
        else:
            total += len(val)
    for key in keys_to_delete:
        del dict[key]
    current = 0
    start_time = time.time()
    time.sleep(1)
    print(f'\n正在通过店铺内商品总销售额筛选并记录到新表')
    for key, val in dict.items():
        for row in val:
            current+=1
            res = (total - current) / (current / ((time.time() - start_time) / 60))
            print(f"\r当前进度：{current}/{total}，预计仍需：{res:.2f} min", end="")
            for cell in sheet[row]:
                new_sheet[f"{get_column_letter(cell.column)}{current+1}"].value = cell.value
except Exception as e:
    print(e)
    print('通过店铺内商品总销售额筛选并记录到新表时出错')

# 处理序号
try:
    start_row = 2
    end_row = new_sheet.max_row

    total = total
    current = 0
    start_time = time.time()
    time.sleep(1)
    print(f'\n正在处理序号')
    for row in range(start_row, end_row + 1):
        current+=1
        res = (total - current) / (current / ((time.time() - start_time) / 60))
        print(f"\r当前进度：{current}/{total}，预计仍需：{res:.2f} min", end="")
        new_sheet.cell(row=row, column=1, value=row-1)
except Exception as e:
    print(e)
    print('处理序号时出错')

new_workbook.save(new_file_name)

# # 修改文件名
# try:
#     temp_file_name = "/".join(file_name.split("/")[:-1]) + '/temp.xlsx'
#     os.rename(file_name, temp_file_name)
#     os.rename(new_file_name, file_name)
#     os.rename(temp_file_name, new_file_name)
# except Exception as e:
#     print(e)
#     print('修改文件名时出错')