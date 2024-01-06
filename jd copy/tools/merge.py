from openpyxl import load_workbook
from openpyxl import Workbook
import time
import os
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles import PatternFill

row_height = 40
column_width = 14

num = 2
new_file_name = 'data/jd/merge/merge.xlsx'
folder_path = 'data/jd/merge'

def merge(new_file_name, folder_path):
    # 新建excel表
    new_workbook = Workbook()
    new_sheet = new_workbook.active

    file_num = 0
    for filename in os.listdir(folder_path):
        if filename.endswith(".xlsx"):
            file_num += 1

    # 处理表头
    headers = ['序号', '电商平台', '关键词/产品', '店铺名称(全称)', '店铺网址', '店铺经营主体信息', '商品图片', '商品标题', '实际品牌', '商品链接', '价格(单位：元)', '销售量(单位：件)', '商品评价(单位：个)', '销售额(单位：元)']
    new_sheet.append(headers)
    new_workbook.save(new_file_name)

    # 记录到新表
    try:
        for index in range(1, file_num+1):
            file_name = os.path.join(folder_path,f'jd ({index}).xlsx')
            # 打开需读取的excel表
            workbook = load_workbook(file_name)
            sheet = workbook.active
            new_workbook = load_workbook(new_file_name)
            new_sheet = new_workbook.active

            start_row = 2
            end_row = sheet.max_row
            last_row = new_sheet.max_row

            total = end_row - start_row + 1
            current = 0
            start_time = time.time()
            time.sleep(1)
            print(f'\n正在记录数据到新表')
            for row in range(start_row, end_row + 1):
                current+=1
                res = (total - current) / (current / ((time.time() - start_time) / 60))
                print(f"\r当前进度：{current}/{total}，预计仍需：{res:.2f} min", end="")
                for cell in sheet[row]:
                    new_sheet[f"{get_column_letter(cell.column)}{last_row+current}"].value = cell.value
            new_workbook.save(new_file_name)
    except Exception as e:
        print(e)
        print('记录到新表时出错')

    # 处理序号
    try:
        start_row = 2
        end_row = new_sheet.max_row

        total = end_row - start_row + 1
        current = 0
        start_time = time.time()
        time.sleep(1)
        print(f'\n正在处理序号')
        for row in range(start_row, end_row + 1):
            current+=1
            res = (total - current) / (current / ((time.time() - start_time) / 60))
            print(f"\r当前进度：{current}/{total}，预计仍需：{res:.2f} min", end="")
            new_sheet.cell(row=row, column=1, value=row-1)
    except Exception as e:
        print(e)
        print('处理序号时出错')
    finally:
        new_workbook.save(new_file_name)