from selenium import webdriver
from openpyxl import Workbook
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import os
import urllib.parse
import time
import random
import datetime
import pygetwindow as gw

min_delay = 3  # 最小延迟时间（单位：秒）
max_delay = 5  # 最大延迟时间（单位：秒）

row_height = 40
column_width = 14

# login函数用于给手动登录操作预留时间
def login(driver):
    print('登录')
    driver.get('https://passport.jd.com/new/login.aspx')
    time.sleep(30)

# scrape_multiple_pages函数用于爬取按照关键词产生的商品从start_page到end_page的分页数据
def scrape_multiple_pages(keyword, start_page, end_page):
    # 打开一个连接固定端口的浏览器模拟器用于保持状态
    options = webdriver.FirefoxOptions()
    driver = webdriver.Remote(
        command_executor="http://127.0.0.1:4444", options=options)
    # options = webdriver.FirefoxOptions()
    # driver = webdriver.Firefox(options=options)

    # 创建带有Selenium Wire的Firefox WebDriver对象
    # options = webdriver.FirefoxOptions()
    # options.set_preference('network.proxy.type', 1)
    # options.set_preference('network.proxy.http', 'localhost')
    # options.set_preference('network.proxy.http_port', 8888)
    # driver = webdriver.Firefox(options=options)

    # login(driver)
    print("登录成功")

    # 获取当前时间并格式化时间字符串，将其作为文件名一部分
    current_time = datetime.datetime.now()
    time_string = current_time.strftime("%Y-%m-%d_%H-%M-%S")
    # 构建文件名，格式为：平台名称+关键词+时间
    file_name = f"data/jd/京东_{urllib.parse.unquote(keyword)}_{time_string}.xlsx"
    # 初始化total_num和record_num用于记录整个过程爬取的商品条数和真正记录到excel表的商品条数
    total_num = 0
    record_num = 0

    # 创建一个新的excel用于记录数据操作并设置好表头为headers的内容，并将得到的新表初步保存一下
    workbook = Workbook()
    sheet = workbook.active
    headers = ['序号', '电商平台', '关键词/产品', '店铺名称(全称)', '店铺网址', '店铺经营主体信息', '商品图片', '商品标题', '实际品牌', '商品链接', '价格(单位：元)', '销售量(单位：件)', '商品评价(单位：个)', '销售额(单位：元)']
    sheet.append(headers)
    workbook.save(file_name)

    # 通过构造url来访问关键词的搜索结果，等待部分时间加载页面，用bs4来解析页面
    url = "https://search.jd.com/Search?keyword="+keyword+"&psort=4&page=1&s=1"
    driver.get(url)
    time.sleep(2)
    html = driver.execute_script(
        "return document.documentElement.outerHTML")
    soup = BeautifulSoup(html, "html.parser")

    # 通过bs4来获取该关键词的搜索得到的商品最大页数，纠正end_page
    try:
        max_page = 1
        verify = soup.select('div.verifyBtn')
        while len(verify) != 0:
            firefox_window = gw.getWindowsWithTitle("Mozilla Firefox")[0]
            firefox_window.minimize()
            firefox_window.maximize()
            firefox_window.activate()
            driver.get(url)
            time.sleep(5)
            html = driver.execute_script(
                "return document.documentElement.outerHTML")
            soup = BeautifulSoup(html, "html.parser")
            verify = soup.select('div.verifyBtn')
            
        elements = soup.select('span.p-skip em b')
        if len(elements):
            max_page = int(elements[0].text)*2
        print("最大页数：",max_page,end_page)
        if(end_page>max_page):
            end_page = max_page      
    except Exception as e:
        print(e)
        print('获取最大页数时出错')

    # 从start_page到end_page爬取商品数据，该过程可以使用ctrl+C来主动终止爬虫程序
    page = start_page
    while page <= end_page:
        try:
            # 不断更新total_num和record_num用于最后能爬取的商品数和实际记录的商品数的统计
            [single_total_num, single_record_num] = scrape_single_page(driver, keyword, page, file_name, headers)
            total_num += single_total_num
            record_num += single_record_num

            if single_total_num == 0:
                firefox_window = gw.getWindowsWithTitle("Mozilla Firefox")[0]
                firefox_window.minimize()
                firefox_window.maximize()
                firefox_window.activate()
                time.sleep(5)
                page-=1
            elif single_total_num != 0 and single_record_num == 0:
                break
        except Exception as e:
            print(e)
            driver.quit()
            print('发生错误，与现有浏览器连接断开')
            break
        except KeyboardInterrupt:
            driver.quit()
            print('用户主动中断爬虫，与现有浏览器连接断开')
            break
        finally:
            page+=1

    driver.quit()
    print('与现有浏览器连接断开')
    # 重命名文件
    new_file_name = f"data/jd/京东_{urllib.parse.unquote(keyword)}_{time_string}_({record_num} of {total_num}).xlsx"
    try:
        os.rename(file_name, new_file_name)
        print(f"已将文件 {file_name} 重命名为 {new_file_name}")
    except Exception as e:
        print(e)
        print(f"重命名文件 {file_name} 失败")
    return [total_num, record_num]

# scrape_single_page函数用于爬取单页的商品信息，并将数据经过部分条件筛选后记录到excel表
def scrape_single_page(driver, keyword, page, file_name, headers):
    # 加载之前创建的excel表
    workbook = load_workbook(file_name)
    sheet = workbook.active
    last_row = sheet.max_row
    # 初始化total_num和record_num用于记录整个过程爬取的商品条数和真正记录到excel表的商品条数
    total_num = 0
    record_num = 0

    print("正在记录第"+str(page)+"页")

    # 通过构造url来访问关键词下搜索得到的每页商品数据
    searchUrl = "https://search.jd.com/Search?keyword=" + \
        keyword+"&psort=4&page="+str(page)+"&s=1"
    driver.get(searchUrl)

    # 设置适当的时间间隔以避免触发反爬虫
    delay = random.uniform(min_delay, max_delay)
    time.sleep(delay)

    html = driver.execute_script(
        "return document.documentElement.outerHTML")
    soup = BeautifulSoup(html, "html.parser")

    # verify = soup.select('div.verifyBtn')
    # while(len(verify)!=0):
    #     driver.get(searchUrl)
    #     html = driver.execute_script(
    #         "return document.documentElement.outerHTML")
    #     soup = BeautifulSoup(html, "html.parser")
    #     time.sleep(2)
    #     verify = soup.select('div.verifyBtn')
        
    # 通过bs4得到整个页面解析得来的单个商品元素构成的列表
    elements = soup.select('li.gl-item')
    try:
        # 逐个解析提取列表中商品信息的各部分数据
        for (index, element) in enumerate(elements, start=1):
            shop_elements = element.select('div.p-shop a.curr-shop.hd-shopname')
            goods_elements = element.select('div.p-img a')
            goods_titles = element.select('div.p-name.p-name-type-2 a em')
            goods_prices = element.select('div.p-price strong i')
            goods_commits = element.select('div.p-commit strong a')

            total_num += 1
            # 筛选
            if (len(shop_elements) != 0):
                if filter_by_shop_name(shop_elements[0].text):
                    continue
            if len(shop_elements) == 0:
                continue
            
            # 筛选掉评论数不足200的商品
            if(len(goods_commits) != 0):
                if convert_string_to_number(goods_commits[0].text and goods_commits[0].text or '0') < 200:
                    continue
            record_num += 1
            
            # 下一行
            last_row+=1
            last_column = 0
            
            # 序号
            try:
                last_column+=1
                ordinal = last_row-1
                sheet.cell(row=last_row, column=last_column, value=ordinal)
            except:
                print(f'记录“{headers[last_column-1]}”时出错')
                return
            
            # 电商平台
            try:
                last_column+=1
                platform_name = '京东'
                sheet.cell(row=last_row, column=last_column, value=platform_name)
            except:
                print(f'记录“{headers[last_column-1]}”时出错')
                return
            
            # 关键词
            try:
                last_column+=1
                search_keyword = urllib.parse.unquote(keyword)
                sheet.cell(row=last_row, column=last_column, value=search_keyword)
            except:
                print(f'记录“{headers[last_column-1]}”时出错')
                return
            
            # 店铺名称
            try:
                last_column+=1
                if (len(shop_elements) != 0):
                    shop_name = shop_elements[0].text
                    sheet.cell(row=last_row, column=last_column, value=shop_name)
                else:
                    sheet.cell(row=last_row, column=last_column, value='')
            except:
                print(f'记录“{headers[last_column-1]}”时出错')
                return
            
            # 店铺网址
            try:
                last_column+=1
                if (len(shop_elements) != 0):
                    shop_link = 'https:' + shop_elements[0].get('href')
                    sheet.cell(row=last_row, column=last_column, value=shop_link)
                else:
                    sheet.cell(row=last_row, column=last_column, value='')
            except:
                print(f'记录“{headers[last_column-1]}”时出错')
                return
            
            # 店铺经营主体信息
            try:
                last_column+=1
            except:
                print(f'记录“{headers[last_column-1]}”时出错')
                return
            
            # 商品图片
            try:
                last_column+=1
                if(len(goods_elements) != 0):
                    goods_img_urls = goods_elements[0].select('img')
                    if (len(goods_img_urls) != 0):
                        goods_img_url = goods_img_urls[0].get('src')
                        if goods_img_url:
                            goods_img_url = 'https:' + goods_img_url
                            if goods_img_url.endswith('.avif'):
                                goods_img_url = goods_img_url[:-5]
                        else:
                            goods_img_url = 'https:' + (goods_elements[0].select('img')[0].get('data-lazy-img'))
                            if goods_img_url.endswith('.avif'):
                                goods_img_url = goods_img_url[:-5]
                        sheet.cell(row=last_row, column=last_column, value=goods_img_url)
            except:
                print(f'记录“{headers[last_column-1]}”时出错')
                return
            
            # 商品标题
            try:
                last_column+=1
                if (len(goods_titles) != 0):
                    goods_title = goods_titles[0].text
                    sheet.cell(row=last_row, column=last_column, value=goods_title)
                else:
                    sheet.cell(row=last_row, column=last_column, value='')
            except:
                print(f'记录“{headers[last_column-1]}”时出错')
                print(goods_titles)
                return
            
            # 商品品牌
            try:
                last_column+=1
            except:
                print(f'记录“{headers[last_column-1]}”时出错')
                return
            
            # 商品链接
            try:
                last_column+=1
                if (len(goods_elements) != 0):
                    goods_link = 'https:' + goods_elements[0].get('href')
                    sheet.cell(row=last_row, column=last_column, value=goods_link)
                else:
                    sheet.cell(row=last_row, column=last_column, value='')
            except:
                print(f'记录“{headers[last_column-1]}”时出错')
                return
            
            # 单价
            try:
                last_column+=1
                if (len(goods_prices) != 0):
                    goods_price = goods_prices[0].text
                    sheet.cell(row=last_row, column=last_column, value=goods_price)
                else:
                    sheet.cell(row=last_row, column=last_column, value='')
            except:
                print(f'记录“{headers[last_column-1]}”时出错')
                return
            
            # 销售量
            try:
                last_column+=1
            except:
                print(f'记录“{headers[last_column-1]}”时出错')
                return
            
            # 商品评论数
            try:
                last_column+=1
                if (len(goods_commits) != 0):
                    goods_commit = goods_commits[0].text and goods_commits[0].text or '0'
                    sheet.cell(row=last_row, column=last_column, value=goods_commit)
                else:
                    sheet.cell(row=last_row, column=last_column, value='0')
            except:
                print(f'记录“{headers[last_column-1]}”时出错')
                return
            
            # 销售额
            try:
                last_column+=1
                if (len(goods_prices) != 0 and len(goods_commits) != 0):
                    goods_price = is_float(goods_prices[0].text) and float(goods_prices[0].text) or 0
                    goods_commit = convert_string_to_number(goods_commits[0].text)
                    goods_sales = goods_price * goods_commit
                    sheet.cell(row=last_row, column=last_column, value=goods_sales)
                else:
                    sheet.cell(row=last_row, column=last_column, value='0')
            except:
                print(f'记录“{headers[last_column-1]}”时出错')
                return
            
    except Exception as e:
        print(e)
        driver.quit()
        print('与现有浏览器连接断开')

    try:
        workbook.save(file_name)
        print(f"已保存第 {page} 页数据到 {file_name}")
    except Exception as e:
        print(e)
        driver.quit()
        print('与现有浏览器连接断开')
    return [total_num, record_num]

# 店铺名称筛选
def filter_by_shop_name(shopName):
    keywords = ['华为京东自营官方旗舰店']
    if shopName in keywords:
        return True
    else:
        return False

# 商品标题筛选
def filter_by_goods_name(goodsName):
    required_keywords = 3  # 至少需要满足的关键字数
    lower_case_good_name = goodsName.lower()
    matched_keywords = 0
    keywords = ['xiaomi', 'huawei', 'oppo', 'vivo', 'redmi', 'realme', '真我', '红米', '小米', '华为', '荣耀', '魅族', '一加', '苹果']
    for keyword in keywords:
        lower_case_keyword = keyword.lower()
        if lower_case_keyword in lower_case_good_name:
            matched_keywords += 1
            if matched_keywords >= required_keywords:
                return True
    return False

# 商品评论数筛选
def filter_by_goods_commit(goodsCommit):
    if not goodsCommit:
        return True
    if goodsCommit.endswith('万+'):
        return False
    elif goodsCommit.endswith('+'):
        if int(goodsCommit[:-1])>=200:
            return False
        else:
            return True
    else:
        return True

# 字符串转数字
def convert_string_to_number(string):
    if not string:
        return 0
    if string.endswith('万+'):
        number = int(string[:-2]) * 10000
    elif string.endswith('+'):
        number = int(string[:-1])
    else:
        number = int(string)
    return number

# 
def is_float(string):
    try:
        float(string)
        return True
    except ValueError:
        return False

def crawl_basic_product_data(keywords, start_page = 1, end_page = 200):
    # keyword为搜索的关键词，依照该关键词到电商平台获取相关商品信息
    # start_page为电商平台分页展示搜索结果，爬取的开始页数
    # end_page为电商平台分页展示搜索结果，爬取的结束页数
    # total_num和record_num分别为整个过程爬取的商品条数和真正记录到excel表的商品条数
    for keyword in  keywords:
        keyword = urllib.parse.quote(keyword)
        [total_num, record_num] = scrape_multiple_pages(keyword, start_page, end_page)
        print(f"共找到 {total_num} 条数据，经过筛选，已记录 {record_num} 条数据")