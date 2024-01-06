# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from openpyxl import Workbook
import time
from openpyxl.utils.cell import get_column_letter
import re
import json

num = 3
recent_json_path = "src/jd/data_files/recent_info/recent_filter.json"
whitelist_json_path = "src/jd/data_files/whitelist_filter.json"
confirm_json_path = "src/jd/data_files/confirm_filter.json"

def filter_by_whitelist(keywords1, keywords2, file_name):
    new_file_name = file_name.replace('.xlsx','_') + str(num) + '.xlsx'
    # 打开需读取的excel表
    workbook = load_workbook(file_name)
    sheet = workbook.active

    # 新建excel表
    new_workbook = Workbook()
    new_sheet = new_workbook.active

    # 处理表头
    print(f'\n正在处理表头')
    first_row = sheet[1]
    for cell in first_row:
        new_sheet[cell.coordinate].value = cell.value

    # 通过店铺白名单筛选
    try:
        def convert_string_to_number(string):
            if not string:
                return 0
            if isinstance(string, int):
                return string
            if string.endswith('万+'):
                number = int(string[:-2]) * 10000
            elif string.endswith('+'):
                number = int(string[:-1])
            else:
                number = int(string)
            return number
        def check_keywords(text, keywords1, keywords2=None):
            pattern1 = '|'.join(keywords1)
            match1 = re.search(pattern1, text, flags=re.IGNORECASE)
            pattern2 = '|'.join(keywords2)
            if keywords2 is not None:
                match2 = re.search(pattern2, text, flags=re.IGNORECASE)
                return match1 is not None and match2 is not None
            return match1 is not None
        
        # 读取白名单店铺信息
        with open(whitelist_json_path, encoding='utf-8') as file:
            white_list = json.load(file)
        # 读取无需筛选店铺信息
        with open(confirm_json_path, encoding='utf-8') as file:
            confirm_list = json.load(file)
        # 读取最近店铺信息
        with open(recent_json_path, encoding='utf-8') as file:
            recent_list = json.load(file)

        record_list = []
        start_row = 2
        end_row = sheet.max_row

        total = end_row - start_row + 1
        current = 0
        start_time = time.time()
        time.sleep(1)
        print(f'\n正在通过店铺白名单筛选')
        for row in range(start_row, end_row + 1):
            shop_name = sheet.cell(row=row, column=4).value
            goods_title = sheet.cell(row=row, column=8).value
            goods_nums = sheet.cell(row=row, column=13).value
            if convert_string_to_number(goods_nums) >= 200 and shop_name not in white_list and shop_name not in confirm_list and shop_name not in recent_list and check_keywords(goods_title, keywords1, keywords2):
                record_list.append(row)
    except Exception as e:
        print(e)
        print('通过店铺白名单筛选时出错')

    # 记录数据到新表
    try:
        start_row = 2
        end_row = sheet.max_row

        total = len(record_list)
        current = 0
        start_time = time.time()
        time.sleep(1)
        print(f'\n正在记录数据到新表')
        for row in record_list:
            current+=1
            res = (total - current) / (current / ((time.time() - start_time) / 60))
            print(f"\r当前进度：{current}/{total}，预计仍需：{res:.2f} min", end="")
            for cell in sheet[row]:
                new_sheet[f"{get_column_letter(cell.column)}{current+1}"].value = cell.value
    except Exception as e:
        print(e)
        print('记录数据到新表时出错')

    # 处理序号
    try:
        start_row = 2
        end_row = new_sheet.max_row

        total = total
        current = 0
        start_time = time.time()
        time.sleep(1)
        print(f'\n正在处理序号')
        for row in range(start_row, end_row + 1):
            current+=1
            res = (total - current) / (current / ((time.time() - start_time) / 60))
            print(f"\r当前进度：{current}/{total}，预计仍需：{res:.2f} min", end="")
            new_sheet.cell(row=row, column=1, value=row-1)
    except Exception as e:
        print(e)
        print('处理序号时出错')

    new_workbook.save(new_file_name)