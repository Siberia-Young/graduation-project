from aip import AipOcr
import ssl
from selenium import webdriver
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from bs4 import BeautifulSoup
import time
from os import path
from aip import AipOcr
import re
import requests
import os
import io
from PIL import Image


""" 你的 APPID AK SK """
APP_ID = '43005493'
API_KEY = 'u6amXwrG9qBDlvjcR7Vo9cVK'
SECRET_KEY = 'x5OjaaiUG441KD6IWkpGaGVApOjQb4lg'

client = AipOcr(APP_ID, API_KEY, SECRET_KEY)
# i = open('OCR/img/code2.png', 'rb')
# img = i.read()
# message = client.webImage(img)
# print(message)

filename = 'due/京东_华为_2023-11-13_17-42-20_(25).xlsx'
row_height = 40
column_width = 14

workbook = load_workbook(filename)
sheet = workbook.active

start_row = 2
end_row = sheet.max_row
aim_col = 5

total = end_row - start_row + 1
current = 0
start_time = time.time()

options = webdriver.FirefoxOptions()
driver = webdriver.Remote(command_executor="http://127.0.0.1:4444", options=options)

# options = webdriver.FirefoxOptions()
# driver = webdriver.Firefox(options=options)

number = []
shop = []


for row in range(start_row, end_row + 1):
    if re.findall(r'\d+', sheet.cell(row=row, column=5).value) not in shop:
        number.append(row)
        shop.append(re.findall(r'\d+', sheet.cell(row=row, column=5).value))

list = [[x, y] for x, y in zip(number, shop)]
time.sleep(1)

total = len(list)
try:
    for element in list:
        current+=1
        res = (total - current) / (current / ((time.time() - start_time) / 60))
        print(f"\r当前进度：{current}/{total}，预计仍需：{res:.2f} min", end="")
        url = 'https://mall.jd.com/showLicence-'+element[1][0]+'.html'
        driver.get(url)
        finish = False
        while not finish:
            tempHTML = driver.execute_script("return document.documentElement.outerHTML")
            tempSoup = BeautifulSoup(tempHTML, "html.parser")
            verifyCodeImg = tempSoup.find_all('img',id='verifyCodeImg')
            if len(verifyCodeImg) != 0:
                try:
                    # image_url = 'https:'+verifyCodeImg[0].get('src')  # 替换为你要识别的图片的URL
                    # print(image_url)
                    # temp_image_path = 'temp/temp.jpg'
                    # response = requests.get(image_url)
                    # with open(temp_image_path, 'wb') as f:
                    #     f.write(response.content)

                    verifyCodeImg = driver.find_element("xpath","//img[@id='verifyCodeImg']")
                    # 获取元素的位置和大小
                    element_location = verifyCodeImg.location
                    element_size = verifyCodeImg.size

                    # 获取整个页面的截图
                    screenshot = driver.get_screenshot_as_png()
                    # 将截图转换为 Image 对象
                    image = Image.open(io.BytesIO(screenshot))

                    # 计算指定元素的区域
                    left = element_location['x'] + 147
                    top = element_location['y'] + 85
                    right = element_location['x'] + element_size['width'] + 173
                    bottom = element_location['y'] + element_size['height'] + 95
                    element_region = (left, top, right, bottom)

                    # 裁剪指定元素的区域
                    element_screenshot = image.crop(element_region)

                    # 保存截图
                    element_screenshot.save('temp/element_screenshot.png')
                    temp_image_path = 'temp/element_screenshot.png'

                    i = open(temp_image_path, 'rb')
                    img = i.read()
                    message = client.webImage(img)
                    if len(message['words_result']) == 0:
                        time.sleep(10)
                        continue
                    code = message['words_result'][0]['words'].replace('-', '').replace(' ', '')
                    

                    verifyCodeInput = driver.find_element("xpath","//input[@id='verifyCode']")
                    verifyCodeInput.send_keys(code)
                    time.sleep(2)
                    sutmit = driver.find_element("xpath","//button[contains(@class, 'btn') and @type='submit']")
                    sutmit.click()
                    time.sleep(1)

                    tempHTML = driver.execute_script("return document.documentElement.outerHTML")
                    tempSoup = BeautifulSoup(tempHTML, "html.parser")
                    qualificationItem = tempSoup.select('li.qualification-item')
                    messageError = tempSoup.select('li.message_error')
                    if len(qualificationItem) != 0:
                        finish = True
                        tempHTML = driver.execute_script("return document.documentElement.outerHTML")
                        tempSoup = BeautifulSoup(tempHTML, "html.parser")
                        elements = tempSoup.select('li.noBorder span')
                        sheet.cell(row=element[0], column=6, value=elements[0].text)
                    elif len(messageError) == 0:
                        finish = True
                        sheet.cell(row=element[0], column=6, value='该店铺无经营企业名称')
                except Exception as e:
                    print(e)
            else:
                finish = True
                sheet.cell(row=element[0], column=6, value='该店铺无证照')
except:
    print('出错')
finally:
    # 保存文件
    workbook.save(filename)
    driver.quit()
    print('与现有浏览器连接断开')
    end_time = time.time()
    duration = end_time - start_time
    print(f"爬虫耗时：{duration:.2f} 秒")
    print(f"目标数量：{total} 条")
    print(f"已获取数量：{current} 条")
    unit = current / (duration / 60)
    print(f"每分钟爬取数量：{unit:.2f} 条")

    

# # 保存文件
# workbook.save(filename)