from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import time
import os
import glob

num = 11

row_height = 40
column_width = 14

def cell_style_adjustments(folder_path):
    xlsx_files = glob.glob(os.path.join(folder_path, "*.xlsx"))
    file_names = [file.replace('\\','/') for file in xlsx_files]
    
    for file_name in file_names:
        print(f'\n正在处理文件：{os.path.basename(file_name)}')
        # 打开需读取的excel表
        workbook = load_workbook(file_name)
        sheet = workbook.active

        # 处理表头
        print(f'\n正在处理表头')
        for index, cell in enumerate(sheet[1], start=1):
            cell.fill = PatternFill(start_color='EB2B2B', end_color='EB2B2B', fill_type='solid')
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 更新单元格样式
        try:
            list = []
            start_row = 2
            end_row = sheet.max_row

            total = end_row - start_row + 1
            current = 0
            start_time = time.time()
            time.sleep(1)
            print(f'\n正在更新单元格样式')
            for row in range(start_row, end_row + 1):
                current+=1
                res = (total - current) / (current / ((time.time() - start_time) / 60))
                print(f"\r当前进度：{current}/{total}，预计仍需：{res:.2f} min", end="")

                # 序号
                sheet.column_dimensions[get_column_letter(1)].width = column_width
                sheet.row_dimensions[row].height = row_height
                ordinal_cell = sheet[f"{get_column_letter(1)}{row}"]
                ordinal_cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
                sheet.cell(row=row, column=1, value=row-1)

                # 电商平台
                sheet.column_dimensions[get_column_letter(2)].width = column_width
                sheet.row_dimensions[row].height = row_height
                platform_name_cell = sheet[f"{get_column_letter(2)}{row}"]
                platform_name_cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')

                # 关键词
                sheet.column_dimensions[get_column_letter(3)].width = column_width
                sheet.row_dimensions[row].height = row_height
                search_keyword_cell = sheet[f"{get_column_letter(3)}{row}"]
                search_keyword_cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')

                # 店铺名称
                sheet.column_dimensions[get_column_letter(4)].width = column_width
                sheet.row_dimensions[row].height = row_height
                shop_name_cell = sheet[f"{get_column_letter(4)}{row}"]
                shop_name_cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')

                # 店铺网址*
                shop_link_url = sheet.cell(row=row, column=5).value
                sheet.column_dimensions[get_column_letter(5)].width = column_width
                sheet.row_dimensions[row].height = row_height
                shop_link_url_cell = sheet[f"{get_column_letter(5)}{row}"]
                shop_link_url_cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
                shop_link_url_cell.font = Font(underline="single", color="0563C1")
                shop_link_url_cell.hyperlink = shop_link_url
                sheet.cell(row=row, column=5, value=shop_link_url)

                # 店铺经营主体信息
                sheet.column_dimensions[get_column_letter(6)].width = column_width
                sheet.row_dimensions[row].height = row_height
                manager_cell = sheet[f"{get_column_letter(6)}{row}"]
                manager_cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')

                # 商品图片*
                goods_img_url = sheet.cell(row=row, column=7).value
                sheet.column_dimensions[get_column_letter(7)].width = column_width
                sheet.row_dimensions[row].height = row_height
                goods_img_url_cell = sheet[f"{get_column_letter(7)}{row}"]
                goods_img_url_cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
                goods_img_url_cell.font = Font(underline="single", color="0563C1")
                goods_img_url_cell.hyperlink = goods_img_url
                sheet.cell(row=row, column=7, value=goods_img_url)

                # 商品标题
                sheet.column_dimensions[get_column_letter(8)].width = column_width
                sheet.row_dimensions[row].height = row_height
                shop_title_cell = sheet[f"{get_column_letter(8)}{row}"]
                shop_title_cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')

                # 商品品牌          
                sheet.column_dimensions[get_column_letter(9)].width = column_width
                sheet.row_dimensions[row].height = row_height
                goods_brand_cell = sheet[f"{get_column_letter(9)}{row}"]
                goods_brand_cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')

                # 商品链接*
                goods_link_url = sheet.cell(row=row, column=10).value
                sheet.column_dimensions[get_column_letter(10)].width = column_width
                sheet.row_dimensions[row].height = row_height
                goods_link_url_cell = sheet[f"{get_column_letter(10)}{row}"]
                goods_link_url_cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
                goods_link_url_cell.font = Font(underline="single", color="0563C1")
                goods_link_url_cell.hyperlink = goods_link_url
                sheet.cell(row=row, column=10, value=goods_link_url)

                # 单价
                sheet.column_dimensions[get_column_letter(11)].width = column_width
                sheet.row_dimensions[row].height = row_height
                goods_price_cell = sheet[f"{get_column_letter(11)}{row}"]
                goods_price_cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')

                # 销售量
                sheet.column_dimensions[get_column_letter(12)].width = column_width
                sheet.row_dimensions[row].height = row_height
                goods_num_cell = sheet[f"{get_column_letter(12)}{row}"]
                goods_num_cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
                            
                # 商品评论数
                sheet.column_dimensions[get_column_letter(13)].width = column_width
                sheet.row_dimensions[row].height = row_height
                goods_commit_cell = sheet[f"{get_column_letter(13)}{row}"]
                goods_commit_cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')

                # 销售额
                sheet.column_dimensions[get_column_letter(14)].width = column_width
                sheet.row_dimensions[row].height = row_height
                goods_sales_cell = sheet[f"{get_column_letter(14)}{row}"]
                goods_sales_cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
        except Exception as e:
            print(e)
            print('更新单元格样式时出错')
        finally:
            # 保存文件
            workbook.save(file_name)
            os.rename(file_name,file_name.replace('.xlsx',f'_({total}条).xlsx'))
