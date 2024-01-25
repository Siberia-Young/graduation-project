from openpyxl import load_workbook
from openpyxl import Workbook
import json
import os
import datetime
import glob

row_height = 40
column_width = 14

def crawl_basic_product_data(keywords, start_page = 1, end_page = 100):
    folder_path = "data/pdd/json"
    json_files = glob.glob(os.path.join(folder_path, "*.json"))
    keywords = [os.path.basename(file).replace('.json','') for file in json_files]
    for keyword in keywords:
        print(f'\n正在处理关键词：{keyword}')
        # 读取JSON文件
        list = []
        with open(f'data/pdd/json/{keyword}.json', 'r', encoding='utf-8') as file:
            list = json.load(file)

        current_time = datetime.datetime.now()
        time_string = current_time.strftime("%Y-%m-%d_%H-%M-%S")
        file_name = f"data/pdd/拼多多_{keyword}_{time_string}.xlsx"
        total_num = 0
        record_num = 0

        workbook = Workbook()
        sheet = workbook.active
        headers = ['序号', '电商平台', '关键词/产品', '店铺名称(全称)', '店铺网址', '店铺经营主体信息', '商品图片', '商品标题', '实际品牌', '商品链接', '价格(单位：元)', '销售量(单位：件)', '商品评价(单位：个)', '销售额(单位：元)']
            
        sheet.append(headers)

        workbook.save(file_name)
        total_num = 0
        record_num = 0
        workbook = load_workbook(file_name)
        sheet = workbook.active
        last_row = sheet.max_row

        def convert_string_to_number(string):
            if not string:
                return 0
            if string.endswith('万+'):
                number = float(string[:-2]) * 10000
            elif string.endswith('万'):
                number = float(string[:-1]) * 10000
            elif string.endswith('+'):
                number = float(string[:-1])
            else:
                number = float(string)
            return number

        for i in range(len(list)):
            for j in range(len(list[i])):
                total_num += 1
                record_num += 1

                # 下一行
                last_row+=1
                last_column = 0

                # 序号
                try:
                    last_column+=1
                    ordinal = last_row-1
                    sheet.cell(row=last_row, column=last_column, value=ordinal)
                except:
                    print(f'记录“{headers[last_column-1]}”时出错')
                    

                # 电商平台
                try:
                    last_column+=1
                    platform_name = '拼多多-批发'
                    sheet.cell(row=last_row, column=last_column, value=platform_name)
                except:
                    print(f'记录“{headers[last_column-1]}”时出错')
                    

                # 关键词
                try:
                    last_column+=1
                    search_keyword = keyword
                    sheet.cell(row=last_row, column=last_column, value=search_keyword)
                except:
                    print(f'记录“{headers[last_column-1]}”时出错')
                    

                # 店铺名称
                try:
                    last_column+=1
                    shop_name = 'mallName' in list[i][j].keys() and list[i][j]['mallName'] or '暂无店铺名称'
                    sheet.cell(row=last_row, column=last_column, value=shop_name)
                except:
                    print(f'记录“{headers[last_column-1]}”时出错')
                    

                # 店铺网址
                try:
                    last_column+=1
                    shop_link = 'mallIdEncrypt' in list[i][j].keys() and ('https://pifa.pinduoduo.com/mall?mid='+list[i][j]['mallIdEncrypt']) or '暂无店铺链接'
                    sheet.cell(row=last_row, column=last_column, value=shop_link)
                except:
                    print(f'记录“{headers[last_column-1]}”时出错')
                    
                
                # 店铺经营主体信息
                try:
                    last_column+=1
                except:
                    print(f'记录“{headers[last_column-1]}”时出错')
                    

                # 商品图片
                try:
                    last_column+=1
                    goods_img_url = 'goodsImgUrl' in list[i][j].keys() and list[i][j]['goodsImgUrl'] or '暂无商品图片'
                    sheet.cell(row=last_row, column=last_column, value=goods_img_url)
                except:
                    print(f'记录“{headers[last_column-1]}”时出错')

                # 商品标题
                try:
                    last_column+=1
                    goods_title = 'goodsName' in list[i][j].keys() and list[i][j]['goodsName'] or '暂无商品标题'
                    sheet.cell(row=last_row, column=last_column, value=goods_title)
                except:
                    print(f'记录“{headers[last_column-1]}”时出错')
                    

                # 商品品牌
                try:
                    last_column+=1
                except:
                    print(f'记录“{headers[last_column-1]}”时出错')
                    

                # 商品链接
                try:
                    last_column+=1
                    goods_link = 'goodsId' in list[i][j].keys() and 'https://pifa.pinduoduo.com/goods/detail/?gid='+str(list[i][j]['goodsId']) or '暂无商品链接'
                    sheet.cell(row=last_row, column=last_column, value=goods_link)
                except:
                    print(f'记录“{headers[last_column-1]}”时出错')
                    

                # 单价
                try:
                    last_column+=1
                    goods_price = 'goodsWholeSalePrice' in list[i][j].keys() and (list[i][j]['goodsWholeSalePrice']/100) or '暂无单价'
                    sheet.cell(row=last_row, column=last_column, value=goods_price)
                except:
                    print(f'记录“{headers[last_column-1]}”时出错')
                    

                # 销售量
                try:
                    last_column+=1
                    goods_num = 'salesTipAmount' in list[i][j].keys() and list[i][j]['salesTipAmount'] or '0'
                    sheet.cell(row=last_row, column=last_column, value=goods_num)
                except:
                    print(f'记录“{headers[last_column-1]}”时出错')
                    
                
                # 商品评论数
                try:
                    last_column+=1
                except:
                    print(f'记录“{headers[last_column-1]}”时出错')
                    
                # 销售额
                try:
                    last_column+=1
                    goods_sales = goods_price * convert_string_to_number(goods_num)
                    sheet.cell(row=last_row, column=last_column, value=goods_sales)
                except:
                    print(f'记录“{headers[last_column-1]}”时出错')
                        
        workbook.save(file_name)