from selenium import webdriver
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import time
import random
from selenium.webdriver.common.proxy import Proxy, ProxyType

from selenium import webdriver
from selenium.webdriver.firefox.options import Options

file_name = "data/tm/需求1.xlsx"
num = 10

# 打开需读取的excel表
workbook = load_workbook(file_name)
sheet = workbook.active

# 创建代理对象
proxy = Proxy()
proxy.proxy_type = ProxyType.MANUAL
proxy.http_proxy = '222.85.6.48:45141'
# 打开火狐浏览器模拟器
options = webdriver.FirefoxOptions()
options.add_argument('--proxy-server={}'.format(proxy.http_proxy))
driver = webdriver.Remote(command_executor="http://127.0.0.1:4444", options=options)

# options = webdriver.EdgeOptions()
# driver = webdriver.Edge(options=options)

# 爬取详细商品数据
try:
    start_row = 2
    end_row = sheet.max_row

    total = end_row - start_row + 1
    current = 0
    start_time = time.time()
    time.sleep(1)
    print(f'\n正在爬取详细商品数据')
    for row in range(start_row, end_row + 1):
        current+=1
        res = (total - current) / (current / ((time.time() - start_time) / 60))
        print(f"\r当前进度：{current}/{total}，预计仍需：{res:.2f} min", end="")
        goods_link = sheet.cell(row=row, column=10).value
        driver.get(goods_link)
        tempHTML = driver.execute_script("return document.documentElement.outerHTML")
        tempSoup = BeautifulSoup(tempHTML, "html.parser")

        elements = tempSoup.select('span[title^="品牌："][class="Attrs--attr--33ShB6X"]')
        print(elements)
        if len(elements) == 0:
            sheet.cell(row=row, column=9, value='暂无')
        else:
            sheet.cell(row=row, column=9, value=elements[0].text.replace('品牌：',''))
except Exception as e:
    print(e)
    print('爬取详细商品数据时出错')
finally:
    # 保存文件
    workbook.save(file_name)
    driver.quit()
    print('与现有浏览器连接断开')
    end_time = time.time()
    duration = end_time - start_time
    print(f"爬虫耗时：{duration:.2f} 秒")
    print(f"目标数量：{total} 条")
    print(f"已获取数量：{current} 条")
    unit = current / (duration / 60)
    print(f"每分钟爬取数量：{unit:.2f} 条")