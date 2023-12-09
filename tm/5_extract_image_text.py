import easyocr
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import time
import os
import re

file_name = "data/tm/需求1_天猫_2023-11-22_18-15-08_(2031).xlsx"
num = 5
match = re.search(r"\d{4}-\d{2}-\d{2}_\d{2}-\d{2}-\d{2}", file_name)
folder_path = "/".join(file_name.split("/")[:-1]) + '/aimages_' + match.group()

# 打开需读取的excel表
workbook = load_workbook(file_name)
sheet = workbook.active

# 只需要运行一次就可以将模型加载到内存中
reader = easyocr.Reader(['ch_sim','en'])

# 识别图片提取文字
try:
    start_row = 2
    end_row = sheet.max_row

    total = end_row - start_row + 1
    current = 0
    start_time = time.time()
    time.sleep(1)
    print(f'\n正在识别图片提取文字')
    for row in range(start_row, end_row + 1):
        current+=1
        res = (total - current) / (current / ((time.time() - start_time) / 60))
        print(f"\r当前进度：{current}/{total}，预计仍需：{res:.2f} min", end="")
        value = sheet.cell(row=row, column=7).value
        if sheet.cell(row=row, column=16).value != None:
            continue
        image_path = os.path.join(folder_path, f'{row}.{value.split(".")[-1]}')
        if os.path.exists(image_path):
            result = reader.readtext(image_path, detail = 0, paragraph=True)
            sheet.cell(row=row, column=16, value=' '.join(result))
except Exception as e:
    print(e)
    print('出错')
finally:
    workbook.save(file_name)
    end_time = time.time()
    duration = end_time - start_time
    print(f"识别耗时：{(duration/60):.2f} min")
    print(f"目标数量：{total} 条")
    print(f"已识别数量：{current} 条")
    unit = current / (duration / 60)
    print(f"每分钟识别数量：{unit:.2f} 条")