from openpyxl import load_workbook
from openpyxl import Workbook
import time
import os
from openpyxl.utils.cell import get_column_letter

file_name = "data/tm/需求2.xlsx"
num = 8
new_file_name = file_name.replace('.xlsx','_') + str(num) + '.xlsx'

# 打开需读取的excel表
workbook = load_workbook(file_name)
sheet = workbook.active

# 新建excel表
new_workbook = Workbook()
new_sheet = new_workbook.active

# 处理表头
print(f'\n正在处理表头')
first_row = sheet[1]
for cell in first_row:
    new_sheet[cell.coordinate].value = cell.value

# 分类并排序
try:
    dict = {}
    start_row = 2
    end_row = sheet.max_row

    total = end_row - start_row + 1
    current = 0
    start_time = time.time()
    time.sleep(1)
    print(f'\n正在分类并排序')
    # 分类
    for row in range(start_row, end_row + 1):
        current+=1
        res = (total - current) / (current / ((time.time() - start_time) / 60))
        print(f"\r当前进度：{current}/{total}，预计仍需：{res:.2f} min", end="")
        value = sheet.cell(row=row, column=5).value
        if value in dict:
            dict[value].append(row)
        else:
            dict[value] = [row]
    # 排序
    def sort_condition(item):
        return int(sheet.cell(row=item, column=14).value)
    for key, val in dict.items():
        val.sort(key=sort_condition, reverse=True)
except Exception as e:
    print(e)
    print('分类并排序时出错')

# 记录到新表
try:
    start_row = 2
    end_row = sheet.max_row

    total = end_row - start_row + 1
    current = 0
    start_time = time.time()
    time.sleep(1)
    print(f'\n正在记录数据到新表')
    for key, val in dict.items():
        for row in val:
            current+=1
            res = (total - current) / (current / ((time.time() - start_time) / 60))
            print(f"\r当前进度：{current}/{total}，预计仍需：{res:.2f} min", end="")
            for cell in sheet[row]:
                new_sheet[f"{get_column_letter(cell.column)}{current+1}"].value = cell.value
except Exception as e:
    print(e)
    print('记录到新表时出错')

# 处理序号
try:
    start_row = 2
    end_row = new_sheet.max_row

    total = end_row - start_row + 1
    current = 0
    start_time = time.time()
    time.sleep(1)
    print(f'\n正在处理序号')
    for row in range(start_row, end_row + 1):
        current+=1
        res = (total - current) / (current / ((time.time() - start_time) / 60))
        print(f"\r当前进度：{current}/{total}，预计仍需：{res:.2f} min", end="")
        new_sheet.cell(row=row, column=1, value=row-1)
except Exception as e:
    print(e)
    print('处理序号时出错')

new_workbook.save(new_file_name)

# 修改文件名
try:
    temp_file_name = "/".join(file_name.split("/")[:-1]) + '/temp.xlsx'
    os.rename(file_name, temp_file_name)
    os.rename(new_file_name, file_name)
    os.rename(temp_file_name, new_file_name)
except Exception as e:
    print(e)
    print('修改文件名时出错')