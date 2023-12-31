from openpyxl import load_workbook
from openpyxl import Workbook
import time
import os
from openpyxl.utils.cell import get_column_letter

file_name = "data/tm/天猫_华为充电宝_2023-11-20_19-26-41_(3982 of 4252).xlsx"
num = 1
base_file_name = 'src/tm/data_files/shop_data.xlsx'
new_file_name = base_file_name.replace('.xlsx','_') + 'new.xlsx'

# 打开需读取的excel表
workbook = load_workbook(file_name)
sheet = workbook.active

workbook = load_workbook(base_file_name)
base_sheet = workbook.active

# 新建excel表
new_workbook = Workbook()
new_sheet = new_workbook.active

# 读取现有店铺信息
try:
    list = []
    record_list = []
    start_row = 2
    end_row = base_sheet.max_row
    for row in range(start_row, end_row + 1):
        value = base_sheet.cell(row=row, column=2).value
        list.append(value)
        record_list.append([base_sheet.cell(row=row, column=1).value,base_sheet.cell(row=row, column=2).value,base_sheet.cell(row=row, column=3).value])
except Exception as e:
    print(e)
    print('读取现有店铺信息时出错')

# 筛选需要新记录的店铺信息
try:  
    start_row = 2
    end_row = sheet.max_row

    total = end_row - start_row + 1
    current = 0
    start_time = time.time()
    time.sleep(1)
    print(f'\n正在筛选需要新记录的店铺信息')
    for row in range(start_row, end_row + 1):
        shop_link = sheet.cell(row=row, column=5).value
        shop_body = sheet.cell(row=row, column=6).value
        if shop_link not in list and shop_body is not None:
            list.append(shop_link)
            record_list.append([sheet.cell(row=row, column=4).value,sheet.cell(row=row, column=5).value,sheet.cell(row=row, column=6).value])
except Exception as e:
    print(e)
    print('筛选需要新记录的店铺信息时出错')

# 记录数据到新表
try:
    start_row = 2
    end_row = len(list) + 1

    total = len(list)
    current = 0
    start_time = time.time()
    time.sleep(1)
    print(f'\n正在记录数据到新表')
    for item in record_list:
        current+=1
        res = (total - current) / (current / ((time.time() - start_time) / 60))
        print(f"\r当前进度：{current}/{total}，预计仍需：{res:.2f} min", end="")
        new_sheet[f"{get_column_letter(1)}{current+1}"].value = item[0]
        new_sheet[f"{get_column_letter(2)}{current+1}"].value = item[1]
        new_sheet[f"{get_column_letter(3)}{current+1}"].value = item[2]
except Exception as e:
    print(e)
    print('记录数据到新表时出错')

new_workbook.save(new_file_name)