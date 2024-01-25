from openpyxl import load_workbook
from openpyxl import Workbook
import time
from openpyxl.utils.cell import get_column_letter

num = 8

def classify_and_sort(file_name):
    new_file_name = file_name.replace('.xlsx','_') + str(num) + '.xlsx'
    # 打开需读取的excel表
    workbook = load_workbook(file_name)
    sheet = workbook.active

    # 新建excel表
    new_workbook = Workbook()
    new_sheet = new_workbook.active

    # 处理表头
    print(f'\n正在处理表头')
    first_row = sheet[1]
    for cell in first_row:
        new_sheet[cell.coordinate].value = cell.value

    # 分类并排序
    try:
        dict = {}
        start_row = 2
        end_row = sheet.max_row

        total = end_row - start_row + 1
        current = 0
        start_time = time.time()
        time.sleep(1)
        print(f'\n正在分类并排序')
        # 分类
        for row in range(start_row, end_row + 1):
            current+=1
            res = (total - current) / (current / ((time.time() - start_time) / 60))
            print(f"\r当前进度：{current}/{total}，预计仍需：{res:.2f} min", end="")
            value = sheet.cell(row=row, column=6).value
            if value in dict:
                dict[value].append(row)
            else:
                dict[value] = [row]
        # 排序
        def sort_condition(item):
            return sheet.cell(row=item, column=14).value
        for key, val in dict.items():
            val.sort(key=sort_condition, reverse=True)
    except Exception as e:
        print(e)
        print('分类并排序时出错')

    # 筛选同商品不同规格的情况
    try:
        def common_prefix_suffix_length(str1, str2):
            prefix_length = 0
            suffix_length = 0
            while prefix_length < len(str1) and prefix_length < len(str2):
                if str1[prefix_length] == str2[prefix_length]:
                    prefix_length += 1
                else:
                    break
            while suffix_length < len(str1) and suffix_length < len(str2):
                if str1[-suffix_length - 1] == str2[-suffix_length - 1]:
                    suffix_length += 1
                else:
                    break
            return [prefix_length+suffix_length, (len(str1) + len(str2)) / 2]
        
        def similarity_rate(base_list, target):
            match_list = []
            for item in base_list:
                if target != '':
                    if target == item:
                        match_list.append(1)
                        break
                    else:
                        [common_length, average_length] = common_prefix_suffix_length(target, item)
                        if common_length == 0:
                            match_list.append(0)
                        else:
                            match_list.append(common_length / average_length)
                else:
                    match_list.append(0)
            return match_list

        def goods_title_match(base_list, target, degree = 0.95):
            contain = False
            index = -1
            match_list = similarity_rate(base_list, target)
            if len(match_list) == len(base_list):
                for key, val in enumerate(match_list):
                    if val >= degree:
                        contain = True
                        if index == -1 or (index != -1 and val > match_list[index]):
                            index = key
            else:
                contain = True
                index = len(match_list) - 1
            if contain:
                return [contain, base_list[index]]
            else:
                return [contain, None]
            
        start_row = 2
        end_row = sheet.max_row

        total = end_row - start_row + 1
        current = 0
        start_time = time.time()
        time.sleep(1)
        print(f'\n正在筛选同商品不同规格的情况')
        for key, val in dict.items():
            new_dict = {}
            for row in val:
                current+=1
                goods_title = sheet.cell(row=row, column=8).value
                [contain, position] = goods_title_match(list(new_dict.keys()), goods_title)
                if contain:
                    new_dict[position].append(row)
                else:
                    new_dict[goods_title] = [row]
            res_list = []
            for new_key, new_val in new_dict.items():
                max_sales = 0
                res = new_val[0]
                for row in new_val:
                    goods_sales = sheet.cell(row=row, column=14).value
                    if goods_sales > max_sales:
                        max_sales = goods_sales
                        res = row
                res_list.append(res)
            dict[key] = res_list
    except Exception as e:
        print(e)
        print('筛选同商品不同规格的情况时出错')

    # 记录到新表
    try:
        start_row = 2
        end_row = sheet.max_row

        total = 0
        for key, val in dict.items():
            total += len(val)
        current = 0
        start_time = time.time()
        time.sleep(1)
        print(f'\n正在记录数据到新表')
        for key, val in dict.items():
            for row in val:
                current+=1
                res = (total - current) / (current / ((time.time() - start_time) / 60))
                print(f"\r当前进度：{current}/{total}，预计仍需：{res:.2f} min", end="")
                for cell in sheet[row]:
                    new_sheet[f"{get_column_letter(cell.column)}{current+1}"].value = cell.value
    except Exception as e:
        print(e)
        print('记录到新表时出错')

    # 处理序号
    try:
        start_row = 2
        end_row = new_sheet.max_row

        total = end_row - start_row + 1
        current = 0
        start_time = time.time()
        time.sleep(1)
        print(f'\n正在处理序号')
        for row in range(start_row, end_row + 1):
            current+=1
            res = (total - current) / (current / ((time.time() - start_time) / 60))
            print(f"\r当前进度：{current}/{total}，预计仍需：{res:.2f} min", end="")
            new_sheet.cell(row=row, column=1, value=row-1)
    except Exception as e:
        print(e)
        print('处理序号时出错')

    new_workbook.save(new_file_name)