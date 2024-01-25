from openpyxl import load_workbook
from openpyxl import Workbook
import time
import os
from openpyxl.utils.cell import get_column_letter
import datetime
import json
import shutil

file_name = "data/jd/2023-12-25/merge/jd_2023-12-25/文件5_(24条).xlsx"
num = 5

base_file_name = "src/jd/data_files/recent_info/recent_filter.json"

try:
    current_time = datetime.datetime.now()
    time_string = current_time.strftime("%Y-%m-%d_%H-%M-%S")
    copy_file_name = base_file_name.replace('.json',f'_{time_string}.json')
    shutil.copy(base_file_name, copy_file_name)
except:
    print(f'\n出错')

# 打开需读取的excel表
workbook = load_workbook(file_name)
sheet = workbook.active

# 记录数据到json
try:
    list = []
    start_row = 2
    end_row = sheet.max_row

    total = end_row - start_row + 1
    current = 0
    start_time = time.time()
    time.sleep(1)
    print(f'\n正在记录数据到json')
    for row in range(start_row, end_row + 1):
        current+=1
        res = (total - current) / (current / ((time.time() - start_time) / 60))
        print(f"\r当前进度：{current}/{total}，预计仍需：{res:.2f} min", end="")
        value = sheet.cell(row=row, column=4).value
        if value not in list:
            list.append(value)
except Exception as e:
    print(e)
finally:
   print(f'\n最近店铺数有：{len(list)}')
   with open(base_file_name, 'w', encoding='utf-8') as file:
        json.dump(list, file, indent=4, ensure_ascii=False)