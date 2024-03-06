from selenium import webdriver
from openpyxl import Workbook
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from io import BytesIO
import os
import requests
import urllib.parse
import time
import random
import datetime
import re
import math

min_delay = 3  # 最小延迟时间（单位：秒）
max_delay = 5  # 最大延迟时间（单位：秒）

row_height = 40
column_width = 14

# login函数用于给手动登录操作预留时间
def login(driver):
    print('登录')
    driver.get('https://login.taobao.com/?redirect_url=https%3A%2F%2Flogin.1688.com%2Fmember%2Fjump.htm%3Ftarget%3Dhttps%253A%252F%252Flogin.1688.com%252Fmember%252FmarketSigninJump.htm%253FDone%253Dhttps%25253A%25252F%25252Fwww.1688.com%25252F&style=tao_custom&from=1688web')
    time.sleep(30)

# scrape_multiple_pages函数用于爬取按照关键词产生的商品从start_page到end_page的分页数据
def scrape_multiple_pages(driver, keyword, start_page, end_page):
    # 获取当前时间并格式化时间字符串，将其作为文件名一部分
    current_time = datetime.datetime.now()
    time_string = current_time.strftime("%Y-%m-%d_%H-%M-%S")
    # 构建文件名，格式为：平台名称+关键词+时间
    file_name = f"data/1688/1688_{urllib.parse.unquote(keyword, encoding='GBK')}_{time_string}.xlsx"
    # 初始化total_num和record_num用于记录整个过程爬取的商品条数和真正记录到excel表的商品条数
    total_num = 0
    record_num = 0

    # 创建一个新的excel用于记录数据操作并设置好表头为headers的内容，并将得到的新表初步保存一下
    workbook = Workbook()
    sheet = workbook.active
    headers = ['序号', '电商平台', '关键词/产品', '店铺名称(全称)', '店铺网址', '店铺经营主体信息', '商品图片', '商品标题', '实际品牌', '商品链接', '价格(单位：元)', '销售量(单位：件)', '商品评价(单位：个)', '销售额(单位：元)']
    sheet.append(headers)
    workbook.save(file_name)

    # 通过构造url来访问关键词的搜索结果，等待部分时间加载页面，用bs4来解析页面
    driver.get("https://s.1688.com/selloffer/offer_search.htm?keywords="+keyword+"&sortType=va_rmdarkgmv30")
    time.sleep(2)
    html = driver.execute_script(
        "return document.documentElement.outerHTML")
    soup = BeautifulSoup(html, "html.parser")

    # 通过bs4来获取该关键词的搜索得到的商品最大页数，纠正end_page
    try:
        max_page_el = soup.select('em.fui-paging-num')
        max_page = 1
        if len(max_page_el) != 0:
            max_page = int(max_page_el[0].text)
        print("最大页数：",max_page,end_page)
        if(end_page>max_page):
            end_page = max_page
    except:
        print('获取最大页数时出错')

    # 从start_page到end_page爬取商品数据，该过程可以使用ctrl+C来主动终止爬虫程序
    for page in range(start_page, end_page+1):
        try:
            tempHtml = driver.execute_script(
                "return document.documentElement.outerHTML")
            # 创建 Beautiful Soup 对象
            tempSoup = BeautifulSoup(tempHtml, "html.parser")
            try:
                max_page_el = tempSoup.select('em.fui-paging-num')
                max_page = 1
                if len(max_page_el) != 0:
                    max_page = int(max_page_el[0].text)
                print('最大页面数：',max_page)
                if(page>max_page):
                    break
            except:
                print('更新最大页数时出错')
            # 不断更新total_num和record_num用于最后能爬取的商品数和实际记录的商品数的统计
            [single_total_num, single_record_num] = scrape_single_page(driver, keyword, start_page, page, file_name, headers)
            total_num += single_total_num
            record_num += single_record_num
            if single_total_num == 0:
                break
        except Exception as e:
            print(e)
            driver.quit()
            print('发生错误，与现有浏览器连接断开')
            break
        except KeyboardInterrupt:
            driver.quit()
            print('用户主动中断爬虫，与现有浏览器连接断开')
            break
        
    # 重命名文件
    new_file_name = f"data/1688/1688_{urllib.parse.unquote(keyword, encoding='GBK')}_{time_string}_({record_num} of {total_num}).xlsx"
    try:
        os.rename(file_name, new_file_name)
        print(f"已将文件 {file_name} 重命名为 {new_file_name}")
    except Exception as e:
        print(e)
        print(f"重命名文件 {file_name} 失败")
    return [total_num, record_num]

# scrape_single_page函数用于爬取单页的商品信息，并将数据经过部分条件筛选后记录到excel表
def scrape_single_page(driver, keyword, start_page, page, file_name, headers):
    # 加载之前创建的excel表
    workbook = load_workbook(file_name)
    sheet = workbook.active
    last_row = sheet.max_row
    # 初始化total_num和record_num用于记录整个过程爬取的商品条数和真正记录到excel表的商品条数
    total_num = 0
    record_num = 0

    # 设置适当的时间间隔以避免触发反爬虫
    delay = random.uniform(min_delay, max_delay)
    time.sleep(delay)

    print("正在记录第"+str(page)+"页")

    # # 通过构造url来访问关键词下搜索得到的每页商品数据
    # searchUrl = "https://s.1688.com/selloffer/offer_search.htm?keywords="+keyword+"&beginPage="+str(page)+"&sortType=va_rmdarkgmv30"
    # driver.get(searchUrl)
    try:
        if(page!=start_page):
            button = driver.find_element("xpath", "//a[contains(@class, 'fui-next')]")
            driver.execute_script("document.querySelector('a.fui-next').style.position = 'relative';")
            driver.execute_script("document.querySelector('a.fui-next').style.zIndex = '99999';")
            button.click()
            time.sleep(2)
    except Exception as e:
        print(e)
        print('点击下一页时出错')
    try:
        # 缓慢下拉页面
        scroll_height = driver.execute_script("return document.body.scrollHeight;")
        current_height = 0
        scroll_speed = 500  # 每次下拉的距离
        while current_height < scroll_height:
            driver.execute_script(f"window.scrollTo(0, {current_height});")
            current_height += scroll_speed
            time.sleep(0.3)  # 等待一段时间，模拟缓慢下拉的效果
            scroll_height = driver.execute_script("return document.body.scrollHeight;")
    except:
        print('下拉获取页面信息时发生错误')

    html = driver.execute_script(
        "return document.documentElement.outerHTML")
    soup = BeautifulSoup(html, "html.parser")

    # 通过bs4得到整个页面解析得来的单个商品元素构成的列表
    elements = soup.select('div.space-offer-card-box')
    print(len(elements))
    try:
        # 逐个解析提取列表中商品信息的各部分数据
        for (index, element) in enumerate(elements, start=1):
            shop_elements = element.find_all('div', class_='company-name', title=True)
            goods_elements = element.select('div.mojar-element-image a')
            goods_titles = element.select('div.mojar-element-title a div.title')
            goods_prices = element.select('div.showPricec div.price')
            goods_sales = element.select('div.sale div.count')

            total_num += 1
            # 筛选
            if (len(shop_elements) != 0):
                if filter_by_shop_name(shop_elements[0].text):
                    continue
            if len(shop_elements) == 0:
                continue
            if(len(goods_prices) == 0):
                continue
            
            # 筛选掉销售额为0的商品
            if(len(goods_sales) != 0):
                if goods_sales[0].text == '':
                    continue
            if(len(goods_prices) != 0):
                if not is_float(goods_prices[0].text):
                    continue
            
            record_num += 1
            
            # 下一行
            last_row+=1
            last_column = 0
            
            # 序号
            try:
                last_column+=1
                ordinal = last_row-1
                sheet.cell(row=last_row, column=last_column, value=ordinal)
            except:
                print(f'记录“{headers[last_column-1]}”时出错')
                return
            
            # 电商平台
            try:
                last_column+=1
                platform_name = '1688'
                sheet.cell(row=last_row, column=last_column, value=platform_name)
            except:
                print(f'记录“{headers[last_column-1]}”时出错')
                return
            
            # 关键词
            try:
                last_column+=1
                search_keyword = urllib.parse.unquote(keyword, encoding='GBK')
            except:
                print(f'记录“{headers[last_column-1]}”时出错')
                return
            
            # 店铺名称
            try:
                last_column+=1
                if (len(shop_elements) != 0):
                    shop_name = shop_elements[0].get('title')
                    sheet.cell(row=last_row, column=last_column, value=shop_name)
                else:
                    sheet.cell(row=last_row, column=last_column, value='')
            except:
                print(f'记录“{headers[last_column-1]}”时出错')
                return
            
            # 店铺网址
            try:
                last_column+=1
                if (len(shop_elements) != 0):
                    shop_link_elements = shop_elements[0].select('a')
                    if (len(shop_link_elements) != 0):
                        shop_link = shop_link_elements[0].get('href')
                        sheet.cell(row=last_row, column=last_column, value=shop_link)
                    else:
                        sheet.cell(row=last_row, column=last_column, value='')
                else:
                    sheet.cell(row=last_row, column=last_column, value='')
            except:
                print(f'记录“{headers[last_column-1]}”时出错')
                return
            
            # 店铺经营主体信息
            try:
                last_column+=1
            except:
                print(f'记录“{headers[last_column-1]}”时出错')
                return
            
            # 商品图片
            try:
                last_column+=1
                if(len(goods_elements) != 0):
                    goods_img_urls = goods_elements[0].select('div.img')
                    if (len(goods_img_urls) != 0):
                        goods_img_url_style = goods_img_urls[0].get('style')
                        if goods_img_url_style:
                            goods_img_url = re.search(r"url\(.*?[\'\"](.*?)['\"]\)", goods_img_url_style)
                            if goods_img_url:
                                goods_img_url = goods_img_url.group(1)
                                sheet.cell(row=last_row, column=last_column, value=goods_img_url.replace("?_=2020",""))
            except:
                print(f'记录“{headers[last_column-1]}”时出错')
                return
            
            # 商品标题
            try:
                last_column+=1
                if (len(goods_titles) != 0):
                    goods_title = goods_titles[0].text
                    sheet.cell(row=last_row, column=last_column, value=goods_title)
                else:
                    sheet.cell(row=last_row, column=last_column, value='')
            except:
                print(f'记录“{headers[last_column-1]}”时出错')
                print(goods_titles)
                return
            
            # 商品品牌
            try:
                last_column+=1
            except:
                print(f'记录“{headers[last_column-1]}”时出错')
                return
            
            # 商品链接
            try:
                last_column+=1
                if (len(goods_elements) != 0):
                    goods_link = goods_elements[0].get('href')
                    sheet.cell(row=last_row, column=last_column, value=goods_link)
                else:
                    sheet.cell(row=last_row, column=last_column, value='')
            except:
                print(f'记录“{headers[last_column-1]}”时出错')
                return
            
            # 单价
            try:
                last_column+=1
                if (len(goods_prices) != 0):
                    goods_price = goods_prices[0].text
                    sheet.cell(row=last_row, column=last_column, value=goods_price)
                else:
                    sheet.cell(row=last_row, column=last_column, value='')
            except:
                print(f'记录“{headers[last_column-1]}”时出错')
                return
            
            # 销售量
            try:
                last_column+=1
                if len(goods_sales) != 0:
                    goods_sales_text = goods_sales[0].text
                    if (goods_sales_text.startswith('成交') and goods_sales_text.endswith('元')):
                        goods_sales_text = goods_sales_text.replace('成交','').replace('元','')
                    else:
                        goods_sales_text = '0'
                    goods_num = convert_string_to_number(goods_sales_text)/float(goods_price)
                    sheet.cell(row=last_row, column=last_column, value=math.ceil(goods_num))
            except Exception as e:
                print(e)
                print(f'记录“{headers[last_column-1]}”时出错')
                return
            
            # 商品评论数
            try:
                last_column+=1
            except:
                print(f'记录“{headers[last_column-1]}”时出错')
                return
            
            # 销售额
            try:
                last_column+=1
                if (len(goods_sales) != 0):
                    goods_sales_text = goods_sales[0].text
                    if (goods_sales_text.startswith('成交') and goods_sales_text.endswith('元')):
                        goods_sales = convert_string_to_number(goods_sales_text.replace('成交','').replace('元',''))
                        sheet.cell(row=last_row, column=last_column, value=goods_sales)
                    else:
                        sheet.cell(row=last_row, column=last_column, value='0')
                else:
                    sheet.cell(row=last_row, column=last_column, value='0')
            except:
                print(f'记录“{headers[last_column-1]}”时出错')
                return
            
    except Exception as e:
        print(e)
        driver.quit()
        print('与现有浏览器连接断开')

    try:
        workbook.save(file_name)
        print(f"已保存第 {page} 页数据到 {file_name}")
    except Exception as e:
        print(e)
        driver.quit()
        print('与现有浏览器连接断开')
    return [total_num, record_num]

# 店铺名称筛选
def filter_by_shop_name(shopName):
    keywords = ['华为京东自营官方旗舰店']
    if shopName in keywords:
        return True
    else:
        return False

# 商品标题筛选
def filter_by_goods_name(goodsName):
    required_keywords = 3  # 至少需要满足的关键字数
    lower_case_good_name = goodsName.lower()
    matched_keywords = 0
    keywords = ['xiaomi', 'huawei', 'oppo', 'vivo', 'redmi', 'realme', '真我', '红米', '小米', '华为', '荣耀', '魅族', '一加', '苹果']
    for keyword in keywords:
        lower_case_keyword = keyword.lower()
        if lower_case_keyword in lower_case_good_name:
            matched_keywords += 1
            if matched_keywords >= required_keywords:
                return True
    return False

# 商品评论数筛选
def filter_by_goods_commit(goodsCommit):
    if not goodsCommit:
        return True
    if goodsCommit.endswith('万+'):
        return False
    elif goodsCommit.endswith('+'):
        if int(goodsCommit[:-1])>=200:
            return False
        else:
            return True
    else:
        return True

# 字符串转数字
def convert_string_to_number(string):
    if not string:
        return 0
    if string.endswith('万+'):
        number = int(string[:-2]) * 10000
    elif string.endswith('+'):
        number = int(string[:-1])
    else:
        number = int(string)
    return number

def is_float(string):
    try:
        float(string)
        return True
    except ValueError:
        return False

def crawl_basic_product_data(keywords, start_page = 1, end_page = 100):
    # keyword为搜索的关键词，依照该关键词到电商平台获取相关商品信息
    # start_page为电商平台分页展示搜索结果，爬取的开始页数
    # end_page为电商平台分页展示搜索结果，爬取的结束页数
    # total_num和record_num分别为整个过程爬取的商品条数和真正记录到excel表的商品条数
    options = webdriver.FirefoxOptions()
    driver = webdriver.Remote(
        command_executor="http://127.0.0.1:4444", options=options)
    # 创建带有Selenium Wire的Firefox WebDriver对象
    # options = webdriver.FirefoxOptions()
    # options.set_preference('network.proxy.type', 1)
    # options.set_preference('network.proxy.http', 'localhost')
    # options.set_preference('network.proxy.http_port', 8888)
    # driver = webdriver.Firefox(options=options)

    # options = webdriver.FirefoxOptions()
    # options.add_argument("--disable-blink-features=AutomationControlled")
    # options.add_argument("--headless")
    # driver = webdriver.Remote(
    #     command_executor="http://127.0.0.1:4444", options=options)
    # driver.maximize_window()
    # driver = webdriver.Firefox()
    # login(driver)
    print("登录成功")
    for keyword in  keywords:
        keyword = urllib.parse.quote(keyword, encoding='GBK')
        [total_num, record_num] = scrape_multiple_pages(driver, keyword, start_page, end_page)
        print(f"共找到 {total_num} 条数据，经过筛选，已记录 {record_num} 条数据")

    driver.quit()
    print('与现有浏览器连接断开')